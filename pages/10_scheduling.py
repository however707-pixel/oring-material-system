import streamlit as st
import pandas as pd
import re
from datetime import date, timedelta
import sys, os
import plotly.express as px

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.shared import inject_css, render_header, render_sidebar

st.set_page_config(page_title="排程系統", page_icon="🗓", layout="wide")
inject_css()
render_header(title="排程系統", subtitle="Production Scheduling System", badge="生管 PC")
render_sidebar()

TODAY      = date.today()
MIN_BUFFER = 10   # 最低齊料緩衝（2個工作週 = 10個工作天）
REF_YEAR   = TODAY.year

# ── 2026 台灣國定假日（週六日已由 weekday() 排除，這裡只需填平日補假）─────────
# 資料來源：勞動部行政院公告，如有異動請自行更新
TAIWAN_HOLIDAYS = {
    # 元旦
    date(2026, 1, 1),   # 元旦
    date(2026, 1, 2),   # 彈性放假
    # 春節
    date(2026, 2, 16),  # 除夕（彈性放假）
    date(2026, 2, 17),  # 春節初一
    date(2026, 2, 18),  # 春節初二
    date(2026, 2, 19),  # 春節初三
    date(2026, 2, 20),  # 春節初四
    # 228 和平紀念日（2/28 週六，3/2 週一補假）
    date(2026, 3, 2),
    # 兒童節（4/4 週六，4/3 週五補假）
    date(2026, 4, 3),
    # 清明節（4/5 週日，4/6 週一補假）
    date(2026, 4, 6),
    # 勞動節
    date(2026, 5, 1),
    # 端午節（農曆5/5 ≈ 6/19 週五）
    date(2026, 6, 19),
    # 中秋節（農曆8/15 ≈ 9/25 週五）
    date(2026, 9, 25),
    # 國慶日（10/10 週六，10/9 週五補假）
    date(2026, 10, 9),
}

def count_workdays(start: date, end: date) -> int:
    """計算 start（不含）到 end（含）之間的工作天數，排除週六日及國定假日。"""
    if not start or not end or start >= end:
        return 0
    count = 0
    cur = start
    while cur < end:
        cur += timedelta(days=1)
        if cur.weekday() < 5 and cur not in TAIWAN_HOLIDAYS:
            count += 1
    return count


# ═══════════════════════════════════════════════════════════════════════════════
# 工具函式
# ═══════════════════════════════════════════════════════════════════════════════

def parse_date_str(s):
    """將 '5/28'、'6/1' 等字串轉為 date，跨年自動判斷。"""
    try:
        m, d = map(int, s.split('/'))
        dt = date(REF_YEAR, m, d)
        # 若日期比今天早超過 6 個月，視為明年
        if (TODAY - dt).days > 180:
            dt = date(REF_YEAR + 1, m, d)
        return dt
    except Exception:
        return None


def parse_ship_date(raw):
    """
    解析 M欄出貨日，回傳 (ship_date: date|None, ship_label: str)
    格式可能為：
      - datetime    → 2026-05-25
      - 文字日期    → 6/15-國智*750，6/29-國智*1000  (取最早日期)
      - 空 / TBD / 試產 / 00:00:00
    """
    if pd.isna(raw):
        return None, ''
    s = str(raw).strip()
    if s in ('', 'nan', 'None', 'TBD', '試產', '00:00:00'):
        return None, s if s not in ('nan', 'None', '00:00:00') else ''

    # 已是 datetime
    if hasattr(raw, 'date'):
        d = raw.date()
        return d, d.strftime('%Y-%m-%d')

    # 嘗試直接 pd.to_datetime
    try:
        d = pd.to_datetime(s).date()
        return d, d.strftime('%Y-%m-%d')
    except Exception:
        pass

    # 從文字中找所有 M/D 或 M月D日 格式，取最早那個
    found = re.findall(r'(\d{1,2})/(\d{1,2})', s)
    dates = []
    for m_str, d_str in found:
        dt = parse_date_str(f'{m_str}/{d_str}')
        if dt:
            dates.append(dt)
    if dates:
        earliest = min(dates)
        return earliest, s   # label 保留原文
    return None, s


def parse_material_status(text):
    """
    解析 L欄進料狀況內容，回傳：
      latest_date  : 所有料號中最晚的預計到料日（None = 無日期資訊）
      delayed_items: [(料號, 預計日, 逾期天數), ...]  已逾期未到
      iqc_items    : [(料號, 最後日期|None), ...]     目前在 IQC
      future_items : [(料號, 預計日, 距今天數), ...]  尚未到但有明確日期
    """
    if pd.isna(text) or not str(text).strip():
        return None, [], [], []

    lines = str(text).strip().split('\n')
    all_dates    = []
    delayed_items = []
    iqc_items    = []
    future_items = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # 料號（* 前面）
        mat_no = line.split('*')[0].strip() if '*' in line else line[:50]

        # 判斷是否在 IQC（行尾或括號內有 IQC）
        is_iqc = bool(re.search(r'IQC', line, re.IGNORECASE))

        # 擷取所有 (M/D) 日期
        date_strs = re.findall(r'\((\d{1,2}/\d{1,2})\)', line)
        mat_dates = [parse_date_str(s) for s in date_strs]
        mat_dates = [d for d in mat_dates if d]

        if mat_dates:
            latest = max(mat_dates)
            all_dates.append(latest)

            if is_iqc:
                iqc_items.append((mat_no, latest))
            elif latest < TODAY:
                days_late = (TODAY - latest).days
                delayed_items.append((mat_no, latest, days_late))
            else:
                days_to   = (latest - TODAY).days
                future_items.append((mat_no, latest, days_to))
        else:
            # 只有 (IQC) 沒有日期
            if is_iqc:
                iqc_items.append((mat_no, None))

    latest_all = max(all_dates) if all_dates else None
    return latest_all, delayed_items, iqc_items, future_items


def classify_assy(val):
    """ASSY / PACKING 欄：回傳 (label, date|None)"""
    if pd.isna(val):
        return '', None
    s = str(val).strip()
    if s.upper() in ('V', 'V-2/5'):
        return '✓', None
    if hasattr(val, 'date'):
        return val.date().strftime('%m/%d'), val.date()
    try:
        d = pd.to_datetime(s).date()
        return d.strftime('%m/%d'), d
    except Exception:
        return s, None


# ═══════════════════════════════════════════════════════════════════════════════
# 解析整份檔案
# ═══════════════════════════════════════════════════════════════════════════════

def parse_file(uploaded):
    df = pd.read_excel(uploaded, sheet_name='LIST', header=0)

    rows = []
    for _, r in df.iterrows():
        wo       = str(r.iloc[1]).strip() if pd.notna(r.iloc[1]) else ''
        product  = str(r.iloc[2]).strip() if pd.notna(r.iloc[2]) else ''
        qty      = r.iloc[3]
        rate     = float(r.iloc[5]) if pd.notna(r.iloc[5]) else 0.0
        hint     = str(r.iloc[8]).strip() if pd.notna(r.iloc[8]) else ''
        need_n   = int(r.iloc[9])  if pd.notna(r.iloc[9])  else 0
        lack_n   = int(r.iloc[10]) if pd.notna(r.iloc[10]) else 0
        mat_text = r.iloc[11]
        delay_mat = str(r.iloc[13]).strip() if pd.notna(r.iloc[13]) else ''

        if not wo or wo == 'nan':
            continue

        ship_date, ship_label = parse_ship_date(r.iloc[12])
        assy_label, assy_date = classify_assy(r.iloc[6])

        latest_mat, delayed, iqc, future = parse_material_status(mat_text)

        # ── 齊料日判斷 ──────────────────────────────────────────────────────
        # 優先用 L欄解析出的最晚到料日；
        # L欄空白時，若 I欄(重點提示)含「已發料」「已齊料」「已發放」→ 視為齊料，
        # 並從 hint 中抽取日期（如「5/4已發料」→ 5/4）
        qi_liao_date = latest_mat

        hint_qi_keywords = ('已發料', '已齊料', '已發放', '齊料')
        hint_is_qi = any(kw in hint for kw in hint_qi_keywords)

        if qi_liao_date is None and hint_is_qi:
            # 從 hint 抽取第一個 M/D 日期（如 5/4、5/13）
            m = re.search(r'(\d{1,2})/(\d{1,2})', hint)
            if m:
                qi_liao_date = parse_date_str(f'{m.group(1)}/{m.group(2)}')
            else:
                qi_liao_date = TODAY   # 有發料字樣但沒日期，保守用今天

        # 料況狀態文字
        if rate >= 1.0 or hint_is_qi:
            mat_status   = '已齊料'
            qi_liao_date = qi_liao_date or TODAY
        elif rate == 0.0:
            mat_status = '完全缺料'
        else:
            mat_status = f'缺料 {rate:.0%}'

        # 齊料緩衝 & 達標差距
        buffer_days = None
        target_gap  = None

        if ship_date and qi_liao_date:
            buffer_days = count_workdays(qi_liao_date, ship_date)
            target_gap  = buffer_days - MIN_BUFFER   # 正=達標, 負=不足

        # 延遲料況摘要（最嚴重的那顆）
        delay_summary = ''
        if delayed:
            worst = max(delayed, key=lambda x: x[2])
            delay_summary = f'逾期 -{worst[2]}天：{worst[0]}'
            if len(delayed) > 1:
                delay_summary += f' 等共 {len(delayed)} 料'
        elif iqc:
            delay_summary = f'IQC 中 {len(iqc)} 料'

        rows.append({
            '工單':         wo,
            '成品料號':     product,
            '預計產量':     qty,
            '出貨日':       ship_date,
            '出貨日_顯示':  ship_label,
            '整體料齊率':   rate,
            '料況狀態':     mat_status,
            'ASSY齊料日':   assy_label,
            'ASSY_date':    assy_date,
            '重點提示':     hint,
            '需領料數':     need_n,
            '未領料數':     lack_n,
            '預計齊料日':   qi_liao_date,
            '緩衝天數':     buffer_days,
            '達標差距':     target_gap,
            '延遲料況':     delay_summary,
            '延遲物料':     str(mat_text) if pd.notna(mat_text) else '',
            '進料明細':     str(mat_text) if pd.notna(mat_text) else '',
            '_delayed':     delayed,
            '_iqc':         iqc,
            '_future':      future,
        })

    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# 說明卡片
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown("#### 系統邏輯說明")

c1, c2, c3, c4 = st.columns(4)
with c1:
    st.markdown("""
<div style="background:#eff6ff;border-left:4px solid #3b82f6;border-radius:6px;padding:12px 14px">
<div style="font-size:13px;font-weight:700;color:#1d4ed8;margin-bottom:6px">📦 出貨日（M欄）</div>
<div style="font-size:12px;color:#374151;line-height:1.6">
以出貨日為中心追蹤每張工單<br>
<b>空白</b> = 尚未回交期給業務
</div>
</div>""", unsafe_allow_html=True)

with c2:
    st.markdown("""
<div style="background:#f0fdf4;border-left:4px solid #22c55e;border-radius:6px;padding:12px 14px">
<div style="font-size:13px;font-weight:700;color:#15803d;margin-bottom:6px">⚙️ 預計齊料日</div>
<div style="font-size:12px;color:#374151;line-height:1.6">
① <b>L欄</b>：取各料號最晚到料日<br>
② L欄空白 → 看 <b>I欄</b> 是否含<br>「已發料／已齊料／已發放」
</div>
</div>""", unsafe_allow_html=True)

with c3:
    st.markdown("""
<div style="background:#fefce8;border-left:4px solid #eab308;border-radius:6px;padding:12px 14px">
<div style="font-size:13px;font-weight:700;color:#854d0e;margin-bottom:6px">📏 緩衝天數</div>
<div style="font-size:12px;color:#374151;line-height:1.6">
<b>出貨日 − 齊料日（工作天）</b><br>
已扣除週六日 + 國定假日<br>
最低標準：<b>≥ 10 工作天（2週）</b><br>
達標差距 = 緩衝 − 10
</div>
</div>""", unsafe_allow_html=True)

with c4:
    st.markdown("""
<div style="background:#fff1f2;border-left:4px solid #ef4444;border-radius:6px;padding:12px 14px">
<div style="font-size:13px;font-weight:700;color:#b91c1c;margin-bottom:6px">🔴 進料延遲判斷</div>
<div style="font-size:12px;color:#374151;line-height:1.6">
L欄格式：<code>料號*數量(日期)</code><br>
日期已過未到 → <b>逾期 −N 天</b><br>
含 IQC → 已到廠驗收中
</div>
</div>""", unsafe_allow_html=True)

st.markdown("""
<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;margin-top:10px;font-size:12px;color:#64748b">
<b>達標差距色碼：</b>
&nbsp;&nbsp;
<span style="background:#dcfce7;color:#15803d;padding:2px 8px;border-radius:4px;font-weight:600">+N 工作天 ✅ 達標（≥10工作天）</span>
&nbsp;
<span style="background:#fef9c3;color:#92400e;padding:2px 8px;border-radius:4px;font-weight:600">−1~−3 工作天 ⚠️ 輕微不足</span>
&nbsp;
<span style="background:#fee2e2;color:#dc2626;padding:2px 8px;border-radius:4px;font-weight:600">< −3 工作天 🔴 嚴重不足</span>
</div>
""", unsafe_allow_html=True)

# ── 自動抓取最新檔案 ──────────────────────────────────────────────────────────
BASE_DIR      = r"\\192.168.2.34\MO_Storage\ORing MO\ORing-MO 工作\早會資料夾"
FILE_NAME     = "簡版-工單缺料狀況.xlsx"
_DATA_DIR     = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
DATA_SCHED_NEW = os.path.join(_DATA_DIR, "kanban_latest.xlsx")
DATA_SCHED_OLD = os.path.join(_DATA_DIR, "kanban_prev.xlsx")

def find_latest_files(n=2):
    """遞迴搜尋 BASE_DIR，回傳最新 n 個 FILE_NAME，依修改時間降冪。"""
    import glob, os
    pattern = os.path.join(BASE_DIR, "**", FILE_NAME)
    files = glob.glob(pattern, recursive=True)
    if not files:
        return []
    files.sort(key=os.path.getmtime, reverse=True)
    result = []
    for f in files[:n]:
        mtime = pd.Timestamp(os.path.getmtime(f), unit='s').tz_localize('UTC').tz_convert('Asia/Taipei')
        result.append((f, mtime))
    return result


# 比對欄位定義：(欄位 index, 顯示名稱)
DIFF_COLS = [(5, "料齊率"), (6, "ASSY日"), (7, "PACKING日"),
             (8, "重點提示"), (11, "進料狀況"), (12, "出貨日")]

def build_diff_map(path_new, path_old):
    """
    比對新舊兩份 LIST，回傳 {工單: [變更欄位名稱, ...]}
    新增工單標記 ['🆕 新增']，消失工單不顯示。
    """
    def load_wo(path):
        df = pd.read_excel(path, sheet_name='LIST', header=0)
        return {str(r.iloc[1]).strip(): r for _, r in df.iterrows()
                if pd.notna(r.iloc[1]) and str(r.iloc[1]).strip() not in ('', 'nan')}

    def extract_ship(raw):
        """從出貨日欄取最早日期，回傳 date 或 None。"""
        if pd.isna(raw): return None
        s = str(raw).strip()
        if s in ('', 'nan', 'None', 'TBD', '試產', '00:00:00'): return None
        if hasattr(raw, 'date'): return raw.date()
        try: return pd.to_datetime(s).date()
        except Exception: pass
        found = re.findall(r'(\d{1,2})/(\d{1,2})', s)
        dates = [parse_date_str(f'{m}/{d}') for m, d in found]
        dates = [d for d in dates if d]
        return min(dates) if dates else None

    def extract_latest_mat_date(text):
        """從 L欄取所有日期中最晚的，回傳 date 或 None。"""
        if pd.isna(text) or not str(text).strip(): return None
        dates = []
        for s in re.findall(r'\((\d{1,2}/\d{1,2})\)', str(text)):
            d = parse_date_str(s)
            if d: dates.append(d)
        return max(dates) if dates else None

    new_map = load_wo(path_new)
    old_map = load_wo(path_old)

    diff = {}
    for wo, r_new in new_map.items():
        if wo not in old_map:
            diff[wo] = ['🆕 新工單']
            continue
        r_old = old_map[wo]
        changed = []

        # ① 出貨日：空→有、提前、延後
        sd_new = extract_ship(r_new.iloc[12] if len(r_new) > 12 else None)
        sd_old = extract_ship(r_old.iloc[12] if len(r_old) > 12 else None)
        if sd_new != sd_old:
            if sd_old is None and sd_new is not None:
                changed.append(f'📅 出貨日新增 →{sd_new.strftime("%m/%d")}')
            elif sd_old is not None and sd_new is None:
                changed.append(f'📅 出貨日移除')
            elif sd_old and sd_new:
                delta = (sd_new - sd_old).days
                arrow = f'提前 {abs(delta)}天' if delta < 0 else f'延後 {delta}天'
                changed.append(f'📅 出貨日{arrow} ({sd_old.strftime("%m/%d")}→{sd_new.strftime("%m/%d")})')

        # ② 進料交期：最晚到料日提早或延後
        md_new = extract_latest_mat_date(r_new.iloc[11] if len(r_new) > 11 else None)
        md_old = extract_latest_mat_date(r_old.iloc[11] if len(r_old) > 11 else None)
        if md_new != md_old:
            if md_old is None and md_new is not None:
                changed.append(f'📦 進料日新增 →{md_new.strftime("%m/%d")}')
            elif md_old is not None and md_new is None:
                changed.append(f'📦 進料日移除')
            elif md_old and md_new:
                delta = (md_new - md_old).days
                arrow = f'提早 {abs(delta)}天' if delta < 0 else f'延後 {delta}天'
                changed.append(f'📦 進料{arrow} ({md_old.strftime("%m/%d")}→{md_new.strftime("%m/%d")})')

        # ③ 料齊率：改善或惡化
        rate_new = float(r_new.iloc[5]) if pd.notna(r_new.iloc[5]) else 0.0
        rate_old = float(r_old.iloc[5]) if pd.notna(r_old.iloc[5]) else 0.0
        if abs(rate_new - rate_old) > 0.001:
            arrow = '↑改善' if rate_new > rate_old else '↓惡化'
            changed.append(f'料齊率{arrow} ({rate_old:.0%}→{rate_new:.0%})')

        # ④ 重點提示變更
        hint_new = str(r_new.iloc[8]).strip() if pd.notna(r_new.iloc[8]) else ''
        hint_old = str(r_old.iloc[8]).strip() if pd.notna(r_old.iloc[8]) else ''
        if hint_new != hint_old and hint_new:
            changed.append(f'💬 {hint_old or "無"}→{hint_new}')

        if changed:
            diff[wo] = changed
    return diff


st.markdown("---")

# ── NAS 偵測 & 自動載入 ────────────────────────────────────────────────────────
all_files    = find_latest_files(2)
nas_ok       = len(all_files) > 0
latest_path  = all_files[0][0] if nas_ok else None
latest_mtime = all_files[0][1] if nas_ok else None
prev_path    = all_files[1][0] if len(all_files) >= 2 else None
prev_mtime   = all_files[1][1] if len(all_files) >= 2 else None

if nas_ok:
    # ── NAS 連線：自動載入，無需按鈕 ────────────────────────────────────────
    rel      = latest_path.replace(BASE_DIR, "").lstrip("\\")
    prev_rel = prev_path.replace(BASE_DIR, "").lstrip("\\") if prev_path else "無"
    st.markdown(
        f"<div style='padding:10px 14px;background:#f0fdf4;border:1px solid #86efac;"
        f"border-radius:6px;font-size:13px'>"
        f"✅ <b>NAS 已連線，自動載入最新版</b>"
        f"&nbsp;&nbsp;📂 {rel}"
        f"&nbsp;&nbsp;<span style='color:#64748b'>（{latest_mtime.strftime('%Y-%m-%d %H:%M')}）</span>"
        f"&nbsp;&nbsp;&nbsp;🔄 比對：{prev_rel}"
        f"</div>",
        unsafe_allow_html=True
    )
    if st.button("🔄 重新偵測 NAS", key="sched_refresh"):
        for k in ("sched_df", "sched_src", "sched_time"):
            st.session_state.pop(k, None)
        st.rerun()

    if "sched_df" not in st.session_state:
        with st.spinner("讀取中..."):
            sdf = parse_file(latest_path)
            if prev_path:
                diff_map = build_diff_map(latest_path, prev_path)
                sdf['變更'] = sdf['工單'].map(
                    lambda wo: '、'.join(diff_map[wo]) if wo in diff_map else '')
            else:
                sdf['變更'] = ''
            st.session_state.sched_df   = sdf
            st.session_state.sched_src  = latest_path
            st.session_state.sched_time = latest_mtime
        changed_n = (sdf['變更'] != '').sum()
        st.success(f"載入完成：{len(sdf)} 張工單，其中 {changed_n} 張與上次不同")

else:
    # ── NAS 離線：先嘗試 data/ 已同步資料，再提示手動上傳 ─────────────────────
    if os.path.exists(DATA_SCHED_NEW) and "sched_df" not in st.session_state:
        with st.spinner("載入已同步的工單資料..."):
            try:
                sdf = parse_file(DATA_SCHED_NEW)
                diff_map = build_diff_map(DATA_SCHED_NEW, DATA_SCHED_OLD) if os.path.exists(DATA_SCHED_OLD) else {}
                sdf["變更"] = sdf["工單"].map(lambda wo: "、".join(diff_map[wo]) if wo in diff_map else "")
                st.session_state.sched_df   = sdf
                st.session_state.sched_src  = "data/kanban_latest.xlsx"
                st.session_state.sched_time = pd.Timestamp(os.path.getmtime(DATA_SCHED_NEW), unit="s")
                changed_n = (sdf["變更"] != "").sum()
                st.success(f"✅ 已從同步資料載入：{len(sdf)} 張工單，其中 {changed_n} 張與上次不同")
            except Exception as e:
                st.warning(f"⚠️ 讀取同步資料失敗：{e}")

    if "sched_df" not in st.session_state:
        st.warning("⚠️ NAS 離線，請手動上傳檔案（雲端使用時）")
        with st.expander("📂 上傳 簡版-工單缺料狀況.xlsx", expanded=True):
            up_col1, up_col2 = st.columns(2)
            with up_col1:
                st.caption("**今日檔案**（必要）")
                upload_new = st.file_uploader(
                    "上傳今日版本", type=["xlsx", "xls"], key="sched_upload_new",
                    label_visibility="collapsed")
            with up_col2:
                st.caption("**前一日檔案**（選填，用於比對變更）")
                upload_old = st.file_uploader(
                    "上傳前一版本", type=["xlsx", "xls"], key="sched_upload_old",
                    label_visibility="collapsed")

        if upload_new is not None:
            if st.button("📥 載入", type="primary", key="sched_load_btn"):
                with st.spinner("讀取中..."):
                    sdf = parse_file(upload_new)
                    if upload_old is not None:
                        diff_map = build_diff_map(upload_new, upload_old)
                        sdf['變更'] = sdf['工單'].map(
                            lambda wo: '、'.join(diff_map[wo]) if wo in diff_map else '')
                    else:
                        sdf['變更'] = ''
                    st.session_state.sched_df   = sdf
                    st.session_state.sched_src  = upload_new.name
                    st.session_state.sched_time = pd.Timestamp.now()
                changed_n = (sdf['變更'] != '').sum()
                st.success(f"載入完成：{len(sdf)} 張工單，其中 {changed_n} 張與上次不同")
        else:
            st.info("👆 請先上傳今日的「簡版-工單缺料狀況.xlsx」")

if "sched_df" not in st.session_state:
    st.stop()

df = st.session_state.sched_df.copy()

st.markdown("---")

# ── 篩選列 ──────────────────────────────────────────────────────────────────
fd1, fd2 = st.columns([3, 4])
with fd1:
    # 出貨日區間：預設今天 ~ 今天 +30 天
    ship_dates = df["出貨日"].dropna()
    d_min = ship_dates.min() if not ship_dates.empty else TODAY
    d_max = ship_dates.max() if not ship_dates.empty else TODAY + timedelta(days=60)
    date_range = st.date_input(
        "出貨日區間",
        value=(TODAY, TODAY + timedelta(days=30)),
        min_value=d_min, max_value=d_max,
        key="ship_range"
    )
with fd2:
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        sel_status = st.selectbox("料況篩選",
            ["全部", "已齊料", "缺料中", "完全缺料", "不達標（<14天）"])
    with fc2:
        sel_delay = st.selectbox("進料延遲",
            ["全部", "有逾期料", "IQC 中", "無問題"])
    with fc3:
        search = st.text_input("工單號 / 料號搜尋", placeholder="輸入關鍵字")

dff = df.copy()

# 出貨日區間篩選（只篩有出貨日的；無出貨日的保留在 Tab2 明細）
if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
    dr_start = pd.Timestamp(date_range[0])
    dr_end   = pd.Timestamp(date_range[1])
    has_date_mask = dff["出貨日"].notna()
    in_range_mask = (dff["出貨日"].apply(lambda v: pd.Timestamp(v) if pd.notna(v) else pd.NaT) >= dr_start) & \
                    (dff["出貨日"].apply(lambda v: pd.Timestamp(v) if pd.notna(v) else pd.NaT) <= dr_end)
    dff = dff[~has_date_mask | in_range_mask]  # 無出貨日的保留；有出貨日的套用區間

if sel_status == "已齊料":
    dff = dff[dff["料況狀態"] == "已齊料"]
elif sel_status == "缺料中":
    dff = dff[dff["料況狀態"] != "已齊料"]
elif sel_status == "完全缺料":
    dff = dff[dff["料況狀態"] == "完全缺料"]
elif sel_status == "不達標（<14天）":
    dff = dff[dff["達標差距"].notna() & (dff["達標差距"] < 0)]

if sel_delay == "有逾期料":
    dff = dff[dff["延遲料況"].str.contains("逾期", na=False)]
elif sel_delay == "IQC 中":
    dff = dff[dff["延遲料況"].str.contains("IQC", na=False)]
elif sel_delay == "無問題":
    dff = dff[~dff["延遲料況"].str.contains("逾期|IQC", na=False)]

if search.strip():
    kw = search.strip()
    dff = dff[dff["工單"].str.contains(kw, na=False) |
              dff["成品料號"].str.contains(kw, na=False)]

# ── KPI（依篩選後結果計算）────────────────────────────────────────────────────
dff_ship   = dff[dff["出貨日"].notna()]
has_ship   = len(dff_ship)
already_qi = (dff_ship["料況狀態"] == "已齊料").sum()
shortage   = has_ship - already_qi

_, k1, k2, k3, _ = st.columns(5)
k1.metric("有出貨日", has_ship)
k2.metric("已齊料",   already_qi)
k3.metric("缺料中",   shortage)
st.markdown("---")

# ═══════════════════════════════════════════════════════════════════════════════
tab0, tab1, tab2, tab3 = st.tabs(["📊 今日出貨看板", "📦 出貨時程總覽", "🔍 進料延遲明細", "📅 每日來料排程"])

# ── Tab0：出貨達成看板 ─────────────────────────────────────────────────────
with tab0:
    import plotly.graph_objects as go

    DAILY_CAP = 200   # 每日產能（pcs）

    # 本週 / 下週 日期範圍（週一～週五）
    _wd     = TODAY.weekday()
    wk_mon  = TODAY - timedelta(days=_wd)
    wk_fri  = wk_mon + timedelta(days=4)
    nwk_mon = wk_mon + timedelta(days=7)
    nwk_fri = nwk_mon + timedelta(days=4)

    def _expand_ship(src_df):
        """
        將出貨日欄含多筆日期的工單（如 '6/12-國智*500，6/18-國智*800'）
        拆成多列，各自帶對應的出貨日與數量。
        沒有多日期的工單維持原樣。
        """
        records = []
        for _, row in src_df.iterrows():
            label = str(row.get("出貨日_顯示", "") or "")
            # 找所有 M/D...*數量 的組合
            pairs = re.findall(r'(\d{1,2}/\d{1,2})[^*\n,，]*\*\s*(\d+)', label)
            if len(pairs) > 1:
                for date_str, qty_str in pairs:
                    d = parse_date_str(date_str)
                    if d:
                        new_row = row.copy()
                        new_row["出貨日"]      = d
                        new_row["出貨日_顯示"] = f"{date_str}*{qty_str}"
                        new_row["預計產量"]    = int(qty_str)
                        records.append(new_row)
            else:
                records.append(row)
        return pd.DataFrame(records) if records else src_df.copy()

    # 展開多日期工單後再使用
    exp_df = _expand_ship(dff[dff["出貨日"].notna()].copy())

    def _week_stats(src_exp, start, end):
        sub = src_exp[(src_exp["出貨日"] >= start) & (src_exp["出貨日"] <= end)].copy()
        total_wo  = len(sub)
        total_qty = int(sub["預計產量"].dropna().sum())
        ready_qty = int(sub[sub["料況狀態"] == "已齊料"]["預計產量"].dropna().sum())
        lack_qty  = total_qty - ready_qty
        ready_wo  = int((sub["料況狀態"] == "已齊料").sum())

        need_days  = round(total_qty / DAILY_CAP, 1) if total_qty else 0
        wdays_left = count_workdays(TODAY - timedelta(days=1), end)
        cap_left   = wdays_left * DAILY_CAP

        # 缺料工單中，最晚的預計齊料日
        lack_sub = sub[sub["料況狀態"] != "已齊料"]
        mat_dates = lack_sub["預計齊料日"].dropna()
        latest_mat = max(mat_dates) if not mat_dates.empty else None

        return dict(total_wo=total_wo, total_qty=total_qty,
                    ready_qty=ready_qty, lack_qty=lack_qty,
                    ready_wo=ready_wo,
                    need_days=need_days, wdays_left=wdays_left,
                    cap_left=cap_left, latest_mat=latest_mat, sub=sub)

    ws  = _week_stats(exp_df, wk_mon,  wk_fri)
    nws = _week_stats(exp_df, nwk_mon, nwk_fri)

    def _render_week_card(label, start, end, stats):
        total_wo   = stats["total_wo"]
        total_qty  = stats["total_qty"]
        ready_qty  = stats["ready_qty"]
        lack_qty   = stats["lack_qty"]
        need_days  = stats["need_days"]
        wdays_left = stats["wdays_left"]
        cap_left   = stats["cap_left"]
        latest_mat = stats["latest_mat"]

        if total_qty == 0:
            verdict = ("⬜", "#f1f5f9", "#64748b", "#e2e8f0", "本週無出貨工單")
        elif lack_qty == 0:
            verdict = ("✅", "#f0fdf4", "#15803d", "#86efac", "全數已齊料，可如期出貨")
        elif cap_left >= lack_qty:
            verdict = ("⚠️", "#fefce8", "#92400e", "#fde68a",
                       f"仍缺料 {lack_qty:,} pcs，但產能尚足（剩餘產能 {cap_left:,} pcs）")
        else:
            verdict = ("🔴", "#fff1f2", "#b91c1c", "#fca5a5",
                       f"缺料 {lack_qty:,} pcs，剩餘產能 {cap_left:,} pcs 不足，有出貨風險")

        icon, bg, text_c, border_c, msg = verdict
        pct_ready = int(ready_qty / total_qty * 100) if total_qty else 0

        mat_line = ""
        if latest_mat is not None:
            mat_str  = latest_mat.strftime('%m/%d') if hasattr(latest_mat, 'strftime') else str(latest_mat)
            mat_line = f"<span style='margin-left:16px'>最晚齊料日 <b style='color:#dc2626'>{mat_str}</b></span>"

        st.markdown(f"""
<div style="border:2px solid {border_c};border-radius:12px;padding:18px 20px;
     background:{bg};height:100%">
  <div style="font-size:18px;font-weight:800;color:{text_c};margin-bottom:2px">
    {icon}&nbsp;{label}
    <span style="font-size:13px;font-weight:400;color:#64748b;margin-left:8px">
      {start.strftime('%m/%d')} ~ {end.strftime('%m/%d')}
    </span>
  </div>
  <div style="font-size:13px;color:{text_c};margin-bottom:12px">{msg}</div>

  <div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:14px">
    <div style="text-align:center">
      <div style="font-size:28px;font-weight:800;color:{text_c}">{total_wo}</div>
      <div style="font-size:11px;color:#64748b">出貨筆數</div>
    </div>
    <div style="text-align:center">
      <div style="font-size:28px;font-weight:800;color:{text_c}">{total_qty:,}</div>
      <div style="font-size:11px;color:#64748b">總數量 (pcs)</div>
    </div>
    <div style="text-align:center">
      <div style="font-size:28px;font-weight:800;color:#15803d">{ready_qty:,}</div>
      <div style="font-size:11px;color:#64748b">已齊料 (pcs)</div>
    </div>
    <div style="text-align:center">
      <div style="font-size:28px;font-weight:800;color:#dc2626">{lack_qty:,}</div>
      <div style="font-size:11px;color:#64748b">缺料 (pcs)</div>
    </div>
    <div style="text-align:center">
      <div style="font-size:28px;font-weight:800;color:{text_c}">{need_days}</div>
      <div style="font-size:11px;color:#64748b">需生產天數<br><span style='font-size:10px'>(200pcs/天)</span></div>
    </div>
  </div>

  <div style="font-size:12px;color:#64748b;margin-bottom:4px">
    齊料進度 {pct_ready}%　｜　週內剩餘產能 {cap_left:,} pcs（{wdays_left} 工作天）{mat_line}
  </div>
  <div style="background:#e2e8f0;border-radius:6px;height:14px;overflow:hidden">
    <div style="display:flex;height:100%">
      <div style="width:{pct_ready}%;background:#22c55e"></div>
      <div style="width:{100-pct_ready}%;background:#f87171"></div>
    </div>
  </div>
</div>""", unsafe_allow_html=True)

    cw1, cw2 = st.columns(2)
    with cw1:
        _render_week_card("本週出貨", wk_mon, wk_fri, ws)
    with cw2:
        _render_week_card("下週出貨", nwk_mon, nwk_fri, nws)

    st.markdown("<div style='margin-top:20px'></div>", unsafe_allow_html=True)

    # ── 缺料工單明細（可展開）────────────────────────────────────────
    both_lack = pd.concat([ws["sub"], nws["sub"]], ignore_index=True)
    both_lack = both_lack[both_lack["料況狀態"] != "已齊料"].drop_duplicates(subset=["工單", "出貨日_顯示"]).sort_values("出貨日")
    if not both_lack.empty:
        with st.expander(f"⚠️ 兩週內缺料工單明細（共 {len(both_lack)} 筆）"):
            show = both_lack[["工單", "成品料號", "預計產量", "出貨日_顯示",
                               "料況狀態", "預計齊料日", "延遲料況", "重點提示"]].copy()
            show["預計齊料日"] = show["預計齊料日"].apply(
                lambda v: v.strftime('%m/%d') if pd.notna(v) and hasattr(v, 'strftime') else "")
            st.dataframe(show, hide_index=True, use_container_width=True)

# ── Tab1：出貨時程總覽 ───────────────────────────────────────────────────────
with tab1:
    st.caption(f"今日：{TODAY}　　最低齊料緩衝：{MIN_BUFFER} 天（2週）　　"
               f"達標差距 = 緩衝天數 - {MIN_BUFFER}，負值代表不達標")

    # 有出貨日的工單，依出貨日排序
    has_date_df = dff[dff["出貨日"].notna()].sort_values("出貨日")
    no_date_df  = dff[dff["出貨日"].isna()]

    def build_display(sub):
        show_cols = ["變更", "工單", "成品料號", "預計產量",
                     "出貨日_顯示", "料況狀態",
                     "預計齊料日", "緩衝天數", "達標差距",
                     "延遲料況", "延遲物料"]
        show_cols = [c for c in show_cols if c in sub.columns]
        return sub[show_cols].copy()

    def highlight_changed_row(row):
        """整列標淡橙底，若該列有變更。"""
        if "變更" in row.index and row["變更"]:
            return ['background-color:#fff7ed'] * len(row)
        return [''] * len(row)

    def color_changed(val):
        if val and str(val).strip():
            return 'background-color:#fed7aa;color:#9a3412;font-weight:bold'
        return ''

    def color_target_gap(val):
        if pd.isna(val):
            return ''
        if val >= 0:
            return 'background-color:#dcfce7;color:#15803d;font-weight:bold'
        elif val >= -3:
            return 'background-color:#fef9c3;color:#92400e;font-weight:bold'
        else:
            return 'background-color:#fee2e2;color:#dc2626;font-weight:bold'

    def color_rate(val):
        if pd.isna(val): return ''
        if val >= 1.0:   return 'color:#15803d;font-weight:bold'
        if val == 0.0:   return 'color:#dc2626;font-weight:bold'
        return 'color:#d97706'

    def color_delay(val):
        if not val: return ''
        if '逾期' in str(val): return 'color:#dc2626;font-weight:bold'
        if 'IQC'  in str(val): return 'color:#d97706'
        return ''

    # ── 匯出 Excel ───────────────────────────────────────────────────────────
    def to_excel_bytes(df_has, df_no):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                     GradientFill)
        from openpyxl.utils import get_column_letter

        EXPORT_COLS = ["變更", "工單", "成品料號", "預計產量", "出貨日_顯示", "料況狀態",
                       "預計齊料日", "緩衝天數", "達標差距", "延遲料況", "延遲物料", "重點提示"]
        COL_WIDTHS  = [20, 22, 36, 10, 16, 12, 12, 10, 10, 30, 52, 16]

        # ── 色彩定義 ────────────────────────────────────────────────────────
        HDR_FILL  = PatternFill("solid", fgColor="1E3A8A")   # 深藍 header
        HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        ROW_EVEN  = PatternFill("solid", fgColor="EFF6FF")   # 淺藍 偶數列
        ROW_ODD   = PatternFill("solid", fgColor="FFFFFF")   # 白色 奇數列

        FILL_GREEN  = PatternFill("solid", fgColor="DCFCE7")
        FILL_YELLOW = PatternFill("solid", fgColor="FEF9C3")
        FILL_RED    = PatternFill("solid", fgColor="FEE2E2")
        FONT_GREEN  = Font(name="Arial", color="15803D", bold=True, size=9)
        FONT_YELLOW = Font(name="Arial", color="92400E", bold=True, size=9)
        FONT_RED    = Font(name="Arial", color="DC2626", bold=True, size=9)
        FONT_ORANGE = Font(name="Arial", color="D97706", size=9)
        FONT_NORMAL = Font(name="Arial", size=9)
        FONT_DELAY  = Font(name="Arial", color="DC2626", size=9)

        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        def prep(sub):
            cols = [c for c in EXPORT_COLS if c in sub.columns]
            out = sub[cols].copy().rename(columns={"出貨日_顯示": "出貨日"})
            if "預計齊料日" in out.columns:
                out["預計齊料日"] = out["預計齊料日"].apply(
                    lambda v: v.strftime('%Y-%m-%d') if pd.notna(v) and hasattr(v, 'strftime')
                    else (str(v) if pd.notna(v) else ""))
            for c in ["緩衝天數", "達標差距"]:
                if c in out.columns:
                    out[c] = out[c].apply(
                        lambda v: f"{v:+.0f}天" if pd.notna(v) else "")
            return out

        def write_sheet(wb, sheet_name, data):
            if data.empty:
                return
            ws = wb.create_sheet(title=sheet_name)
            out = prep(data)
            headers = list(out.columns)

            # ── 標題列 ──────────────────────────────────────────────────────
            ws.row_dimensions[1].height = 22
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill   = HDR_FILL
                cell.font   = HDR_FONT
                cell.alignment = center
                cell.border = border

            # ── 資料列 ──────────────────────────────────────────────────────
            gap_col     = headers.index("達標差距") + 1 if "達標差距" in headers else None
            delay_col   = headers.index("延遲料況") + 1 if "延遲料況" in headers else None
            status_col  = headers.index("料況狀態") + 1 if "料況狀態" in headers else None
            changed_col = headers.index("變更")     + 1 if "變更"     in headers else None
            FILL_CHANGED     = PatternFill("solid", fgColor="FFF7ED")
            FILL_CHANGED_TAG = PatternFill("solid", fgColor="FED7AA")
            FONT_CHANGED_TAG = Font(name="Arial", color="9A3412", bold=True, size=9)

            for ri, (_, row) in enumerate(out.iterrows(), 2):
                ws.row_dimensions[ri].height = 30
                is_even   = (ri % 2 == 0)
                is_changed = bool(str(row.iloc[changed_col - 1]).strip()) if changed_col else False
                row_fill  = FILL_CHANGED if is_changed else (ROW_EVEN if is_even else ROW_ODD)

                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=str(val) if pd.notna(val) else "")
                    cell.border    = border
                    cell.font      = FONT_NORMAL
                    cell.alignment = left

                    # 達標差距 上色
                    if ci == gap_col:
                        s = str(val)
                        if s.startswith("+") or (s and s[0].isdigit()):
                            cell.fill = FILL_GREEN;  cell.font = FONT_GREEN
                        elif s.startswith("-"):
                            days = int(s.replace("天","").replace("+",""))
                            if days >= -3:
                                cell.fill = FILL_YELLOW; cell.font = FONT_YELLOW
                            else:
                                cell.fill = FILL_RED;    cell.font = FONT_RED
                        else:
                            cell.fill = row_fill
                        cell.alignment = center

                    # 延遲料況 上色
                    elif ci == delay_col:
                        cell.fill = row_fill
                        if "逾期" in str(val):
                            cell.font = FONT_DELAY
                        elif "IQC" in str(val):
                            cell.font = FONT_ORANGE
                    # 料況狀態 上色
                    elif ci == status_col:
                        cell.fill = row_fill
                        if "已齊料" in str(val):
                            cell.font = FONT_GREEN
                        elif "完全缺料" in str(val):
                            cell.font = FONT_RED
                        elif "缺料" in str(val):
                            cell.font = FONT_ORANGE
                    # 變更欄 上色
                    elif ci == changed_col:
                        if str(val).strip():
                            cell.fill = FILL_CHANGED_TAG
                            cell.font = FONT_CHANGED_TAG
                            cell.alignment = center
                        else:
                            cell.fill = row_fill
                    else:
                        cell.fill = row_fill

            # ── 欄寬 ────────────────────────────────────────────────────────
            for ci, h in enumerate(headers, 1):
                idx = EXPORT_COLS.index(h.replace("出貨日", "出貨日_顯示")
                                          if h == "出貨日" else h) if (
                    h.replace("出貨日", "出貨日_顯示") if h == "出貨日" else h
                ) in EXPORT_COLS else -1
                ws.column_dimensions[get_column_letter(ci)].width = (
                    COL_WIDTHS[idx] if 0 <= idx < len(COL_WIDTHS) else 15)

            # ── 凍結首列 ────────────────────────────────────────────────────
            ws.freeze_panes = "A2"

        wb = Workbook()
        wb.remove(wb.active)   # 移除預設空白頁
        write_sheet(wb, "有出貨日", df_has)
        write_sheet(wb, "出貨日未定", df_no)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    ex_col, _ = st.columns([1, 5])
    with ex_col:
        if not dff.empty:
            st.download_button(
                "⬇ 匯出 Excel",
                data=to_excel_bytes(has_date_df, no_date_df),
                file_name=f"排程_缺料狀況_{TODAY}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    if has_date_df.empty:
        st.info("篩選條件下無有出貨日的工單。")
    else:
        st.markdown(f"**有出貨日 — {len(has_date_df)} 張**")
        disp = build_display(has_date_df)
        style = (
            disp.style
                .apply(highlight_changed_row, axis=1)
                .map(color_changed,    subset=["變更"]      if "變更"      in disp.columns else [])
                .map(color_target_gap, subset=["達標差距"]  if "達標差距"  in disp.columns else [])
                .map(color_delay,      subset=["延遲料況"]  if "延遲料況"  in disp.columns else [])
                .format({"緩衝天數": lambda v: f"{v:+.0f}天" if pd.notna(v) else "",
                         "達標差距": lambda v: f"{v:+.0f}天" if pd.notna(v) else "",
                         "預計齊料日": lambda v: v.strftime('%m/%d') if pd.notna(v) and hasattr(v, 'strftime') else (str(v) if pd.notna(v) else "")})
        )
        st.dataframe(style, use_container_width=True, hide_index=True)

    if not no_date_df.empty:
        with st.expander(f"⚠️ 出貨日未定 — {len(no_date_df)} 張"):
            disp2 = build_display(no_date_df)
            st.dataframe(disp2.style
                .map(color_rate,  subset=["整體料齊率"] if "整體料齊率" in disp2.columns else [])
                .map(color_delay, subset=["延遲料況"]   if "延遲料況"   in disp2.columns else [])
                .format({"整體料齊率": "{:.0%}"})
                , use_container_width=True, hide_index=True)

# ── Tab2：進料延遲明細 ────────────────────────────────────────────────────────
with tab2:
    st.caption("展開每張工單，逐料顯示進料狀況、是否逾期、逾期天數。")

    # 只顯示有缺料的工單
    need_track = dff[dff["整體料齊率"] < 1.0].sort_values("出貨日")

    if need_track.empty:
        st.success("目前篩選範圍內所有工單均已齊料。")
    else:
        for _, row in need_track.iterrows():
            delayed  = row["_delayed"]   # [(料號, 日期, 逾期天)]
            iqc      = row["_iqc"]       # [(料號, 日期|None)]
            future   = row["_future"]    # [(料號, 日期, 距今天)]

            ship_str = row["出貨日_顯示"] or "出貨日未定"
            gap_str  = (f"達標差距 **{row['達標差距']:+.0f} 天**"
                        if pd.notna(row["達標差距"]) else "出貨日未定")
            warn = "🔴" if (pd.notna(row["達標差距"]) and row["達標差距"] < 0) else (
                   "🟡" if row["延遲料況"] else "🟢")

            with st.expander(
                f"{warn} {row['工單']}　｜　{row['成品料號']}　×{row['預計產量']}　"
                f"｜　出貨 {ship_str}　｜　{gap_str}　"
                f"｜　料齊率 {row['整體料齊率']:.0%}　{row['重點提示']}"
            ):
                c1, c2 = st.columns([1, 2])
                with c1:
                    st.metric("整體料齊率",  f"{row['整體料齊率']:.0%}")
                    st.metric("需/未領料數", f"{row['需領料數']} / {row['未領料數']}")
                    if pd.notna(row["預計齊料日"]):
                        st.metric("預計齊料日", row["預計齊料日"].strftime('%m/%d'))
                    if pd.notna(row["達標差距"]):
                        color = "normal" if row["達標差距"] >= 0 else "inverse"
                        st.metric("達標差距（緩衝-10工作天）",
                                  f"{row['達標差距']:+.0f} 天", delta_color=color)
                    if row["延遲物料"]:
                        st.markdown(f"**N欄延遲物料：** {row['延遲物料']}")

                with c2:
                    if delayed:
                        st.markdown("**🔴 逾期未到（進料日已過但尚未到）**")
                        rows_d = [{"料號": m, "承諾到料日": d.strftime('%m/%d'),
                                   "已逾期": f"-{n} 天"} for m, d, n in
                                  sorted(delayed, key=lambda x: -x[2])]
                        st.dataframe(pd.DataFrame(rows_d), hide_index=True,
                                     use_container_width=True)

                    if iqc:
                        st.markdown("**🟡 IQC 檢驗中**")
                        rows_i = [{"料號": m,
                                   "到廠日": d.strftime('%m/%d') if d else "未知"}
                                  for m, d in iqc]
                        st.dataframe(pd.DataFrame(rows_i), hide_index=True,
                                     use_container_width=True)

                    if future:
                        st.markdown("**🔵 預計進料（未到但有承諾日）**")
                        rows_f = [{"料號": m, "預計到料": d.strftime('%m/%d'),
                                   "距今": f"+{n} 天"} for m, d, n in
                                  sorted(future, key=lambda x: x[1])]
                        st.dataframe(pd.DataFrame(rows_f), hide_index=True,
                                     use_container_width=True)

                    if not delayed and not iqc and not future:
                        st.info("L欄無詳細進料資訊，請參考重點提示欄。")
                        if row["進料明細"]:
                            st.text(row["進料明細"])

# ── Tab3：每日來料排程 ─────────────────────────────────────────────────────
with tab3:
    st.caption("整合所有缺料工單的進料資訊，依預計到料日分組，方便追蹤每天應到的物料。")

    # ── 建立「每日來料」資料表 ──────────────────────────────────────
    # 從每張工單的 _delayed / _future / _iqc 展開成逐料列
    daily_rows = []
    for _, row in dff.iterrows():
        if row["料況狀態"] == "已齊料":
            continue
        wo       = row["工單"]
        product  = row["成品料號"]
        ship_d   = row["出貨日"]
        ship_str = row["出貨日_顯示"] or "未定"
        qty      = row["預計產量"]

        # 計算距出貨日剩餘工作天，判斷是否為急件
        if pd.notna(ship_d):
            wdays_to_ship = count_workdays(TODAY - timedelta(days=1), ship_d)
            is_urgent = wdays_to_ship <= 10
        else:
            wdays_to_ship = None
            is_urgent = False
        urgent_label = "🚨 急件" if is_urgent else "📦 不急件"
        urgent_note  = f"（出貨剩 {wdays_to_ship} 工作天）" if wdays_to_ship is not None else ""

        # 逾期未到
        for mat, arr_d, days_late in row["_delayed"]:
            daily_rows.append({
                "預計到料日":  arr_d,
                "狀態":       "🔴 逾期",
                "急件":        urgent_label,
                "料號":        mat,
                "工單":        wo,
                "成品料號":    product,
                "出貨日":      ship_str,
                "出貨數量":    qty,
                "備註":        f"已逾期 {days_late} 天 {urgent_note}",
                "_urgent":    is_urgent,
            })
        # IQC 中
        for mat, arr_d in row["_iqc"]:
            daily_rows.append({
                "預計到料日":  arr_d if arr_d else None,
                "狀態":       "🟡 IQC",
                "急件":        urgent_label,
                "料號":        mat,
                "工單":        wo,
                "成品料號":    product,
                "出貨日":      ship_str,
                "出貨數量":    qty,
                "備註":        f"廠內驗收中 {urgent_note}",
                "_urgent":    is_urgent,
            })
        # 未來預計到料
        for mat, arr_d, days_to in row["_future"]:
            daily_rows.append({
                "預計到料日":  arr_d,
                "狀態":       "🔵 待進料",
                "急件":        urgent_label,
                "料號":        mat,
                "工單":        wo,
                "成品料號":    product,
                "出貨日":      ship_str,
                "出貨數量":    qty,
                "備註":        f"距今 +{days_to} 天 {urgent_note}",
                "_urgent":    is_urgent,
            })

    if not daily_rows:
        st.success("目前篩選範圍內所有缺料工單均無進料明細資訊（請確認 L欄是否填寫）。")
    else:
        mat_df = pd.DataFrame(daily_rows)

        # ── 篩選列 ───────────────────────────────────────────────────
        fc1, fc2, fc3, fc4 = st.columns([2, 2, 2, 2])
        with fc1:
            _dates = mat_df["預計到料日"].dropna()
            d_min  = _dates.min() if not _dates.empty else TODAY
            d_max  = _dates.max() if not _dates.empty else TODAY + timedelta(days=30)
            mat_range = st.date_input(
                "到料日區間",
                value=(TODAY, min(d_max, TODAY + timedelta(days=30))),
                min_value=d_min, max_value=d_max,
                key="mat_range"
            )
        with fc2:
            status_opts = ["全部"] + sorted(mat_df["狀態"].unique().tolist())
            sel_mat_status = st.selectbox("進料狀態", status_opts, key="mat_status")
        with fc3:
            sel_urgent = st.selectbox("急件篩選",
                ["全部", "🚨 急件（≤10工作天）", "📦 不急件"], key="mat_urgent")
        with fc4:
            mat_search = st.text_input("料號 / 工單搜尋", placeholder="輸入關鍵字", key="mat_kw")

        mdf = mat_df.copy()
        if isinstance(mat_range, (list, tuple)) and len(mat_range) == 2:
            r0, r1 = mat_range
            mdf = mdf[mdf["預計到料日"].isna() |
                      ((mdf["預計到料日"] >= r0) & (mdf["預計到料日"] <= r1))]
        if sel_mat_status != "全部":
            mdf = mdf[mdf["狀態"] == sel_mat_status]
        if sel_urgent == "🚨 急件（≤10工作天）":
            mdf = mdf[mdf["_urgent"] == True]
        elif sel_urgent == "📦 不急件":
            mdf = mdf[mdf["_urgent"] == False]
        if mat_search.strip():
            kw = mat_search.strip()
            mdf = mdf[mdf["料號"].str.contains(kw, na=False) |
                      mdf["工單"].str.contains(kw, na=False)]

        # ── 匯出 Excel ───────────────────────────────────────────────
        def _mat_to_excel(df_export):
            import io
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            COLS   = ["預計到料日", "狀態", "急件", "料號", "工單", "成品料號",
                      "出貨日", "出貨數量", "備註"]
            WIDTHS = [14,           12,     12,     36,    18,    32,
                      14,           10,     36]

            thin   = Side(style="thin", color="CBD5E1")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center")
            left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

            HDR_FILL   = PatternFill("solid", fgColor="1E3A8A")
            HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            FILL_URG   = PatternFill("solid", fgColor="FFF7ED")   # 急件
            FILL_OVD   = PatternFill("solid", fgColor="FEE2E2")   # 逾期
            FILL_IQC   = PatternFill("solid", fgColor="FEF9C3")   # IQC
            FILL_EVEN  = PatternFill("solid", fgColor="EFF6FF")
            FILL_ODD   = PatternFill("solid", fgColor="FFFFFF")
            FONT_NORM  = Font(name="Arial", size=9)
            FONT_URG   = Font(name="Arial", size=9, bold=True, color="C2410C")
            FONT_OVD   = Font(name="Arial", size=9, bold=True, color="DC2626")

            out = df_export[COLS].copy()
            out["預計到料日"] = out["預計到料日"].apply(
                lambda v: v.strftime('%Y-%m-%d') if pd.notna(v) and hasattr(v,'strftime') else "")
            out["出貨數量"] = out["出貨數量"].apply(
                lambda v: int(v) if pd.notna(v) else "")

            wb = Workbook()
            ws = wb.active
            ws.title = f"來料排程_{TODAY.strftime('%m%d')}"
            ws.row_dimensions[1].height = 22

            for ci, h in enumerate(COLS, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = HDR_FILL; cell.font = HDR_FONT
                cell.alignment = center; cell.border = border
                ws.column_dimensions[get_column_letter(ci)].width = WIDTHS[ci-1]

            ws.freeze_panes = "A2"

            for ri, (_, row) in enumerate(out.iterrows(), 2):
                ws.row_dimensions[ri].height = 22
                is_urg = bool(df_export.at[row.name, "_urgent"]) if row.name in df_export.index else False
                status = str(row["狀態"])
                base_fill = (FILL_OVD if "逾期" in status else
                             FILL_IQC if "IQC" in status else
                             FILL_URG if is_urg else
                             (FILL_EVEN if ri % 2 == 0 else FILL_ODD))
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=str(val) if pd.notna(val) else "")
                    cell.border = border
                    cell.fill   = base_fill
                    cell.alignment = center if ci in (1, 7, 8) else left
                    cell.font = (FONT_URG if is_urg and "逾期" not in status else
                                 FONT_OVD if "逾期" in status else FONT_NORM)

            buf = io.BytesIO(); wb.save(buf)
            return buf.getvalue()

        ex1, _ = st.columns([1, 5])
        with ex1:
            st.download_button(
                "⬇ 匯出 Excel",
                data=_mat_to_excel(mdf),
                file_name=f"每日來料排程_{TODAY}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="mat_export"
            )

        # ── KPI ─────────────────────────────────────────────────────
        n_urgent  = mdf["_urgent"].sum()
        n_normal  = (~mdf["_urgent"]).sum()
        n_overdue = (mdf["狀態"] == "🔴 逾期").sum()
        n_iqc     = (mdf["狀態"] == "🟡 IQC").sum()
        n_future  = (mdf["狀態"] == "🔵 待進料").sum()

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("追蹤料號總數", len(mdf))
        k2.metric("🚨 急件",      n_urgent,  delta="需優先追蹤" if n_urgent else None,
                  delta_color="inverse")
        k3.metric("📦 不急件",    n_normal)
        k4.metric("🔴 逾期未到",  n_overdue)
        k5.metric("🟡 IQC 驗收中", n_iqc)
        st.markdown("---")

        # ── 顯示模式切換 ─────────────────────────────────────────────
        view_mode = st.radio(
            "顯示方式",
            ["依到料日分組", "急件優先（上方）", "急件 / 不急件 分開"],
            horizontal=True, key="mat_view"
        )

        wday_names = ["一", "二", "三", "四", "五", "六", "日"]

        def _date_header(arr_date, n_items, extra_tags=""):
            is_past  = arr_date < TODAY
            is_today = arr_date == TODAY
            if is_past:
                hdr_bg, hdr_c = "#fff1f2", "#b91c1c"
                tag = "<span style='background:#fee2e2;color:#b91c1c;border-radius:4px;padding:1px 8px;font-size:12px;margin-left:6px'>⚠️ 已逾期</span>"
            elif is_today:
                hdr_bg, hdr_c = "#eff6ff", "#1d4ed8"
                tag = "<span style='background:#dbeafe;color:#1d4ed8;border-radius:4px;padding:1px 8px;font-size:12px;margin-left:6px'>📌 今日</span>"
            else:
                hdr_bg, hdr_c = "#f8fafc", "#334155"
                tag = ""
            wday = wday_names[arr_date.weekday()]
            st.markdown(f"""
<div style="background:{hdr_bg};border-left:4px solid {hdr_c};border-radius:6px;
     padding:7px 14px;margin-bottom:4px;margin-top:8px">
  <span style="font-size:14px;font-weight:800;color:{hdr_c}">
    {arr_date.strftime('%m / %d')}（週{wday}）
  </span>
  <span style="font-size:12px;color:#64748b;margin-left:8px">{n_items} 項</span>
  {tag}{extra_tags}
</div>""", unsafe_allow_html=True)

        def _show_table(sub_df):
            display = sub_df[["狀態", "急件", "料號", "工單", "成品料號",
                               "出貨日", "出貨數量", "備註"]].copy()
            display["出貨數量"] = display["出貨數量"].apply(
                lambda v: int(v) if pd.notna(v) else "")

            # urgent_map: index → bool，用來配色
            urgent_map = sub_df["_urgent"].to_dict()

            def _rc(r):
                is_urg = urgent_map.get(r.name, False)
                if r["狀態"] == "🔴 逾期":
                    return ["background-color:#fff1f2"] * len(r)
                if r["狀態"] == "🟡 IQC":
                    return ["background-color:#fefce8"] * len(r)
                if is_urg:
                    return ["background-color:#fff7ed"] * len(r)
                return [""] * len(r)

            st.dataframe(
                display.style.apply(_rc, axis=1),
                hide_index=True, use_container_width=True
            )

        has_date = mdf[mdf["預計到料日"].notna()].copy()
        no_date  = mdf[mdf["預計到料日"].isna()].copy()

        if view_mode == "依到料日分組":
            for arr_date in sorted(has_date["預計到料日"].unique()):
                day_df = has_date[has_date["預計到料日"] == arr_date]
                n_urg  = day_df["_urgent"].sum()
                extra  = (f"<span style='background:#fff7ed;color:#c2410c;border-radius:4px;"
                          f"padding:1px 8px;font-size:12px;margin-left:6px'>"
                          f"🚨 急件 {n_urg} 項</span>") if n_urg else ""
                _date_header(arr_date, len(day_df), extra)
                _show_table(day_df)

        elif view_mode == "急件優先（上方）":
            urgent_df = has_date[has_date["_urgent"]].sort_values("預計到料日")
            normal_df = has_date[~has_date["_urgent"]].sort_values("預計到料日")

            if not urgent_df.empty:
                st.markdown("""
<div style="background:#fff7ed;border:2px solid #f97316;border-radius:8px;
     padding:8px 14px;margin-bottom:8px;font-size:15px;font-weight:800;color:#c2410c">
  🚨 急件（出貨日 ≤ 10 工作天）
</div>""", unsafe_allow_html=True)
                for arr_date in sorted(urgent_df["預計到料日"].unique()):
                    day_df = urgent_df[urgent_df["預計到料日"] == arr_date]
                    _date_header(arr_date, len(day_df))
                    _show_table(day_df)

            if not normal_df.empty:
                st.markdown("""
<div style="background:#f0fdf4;border:2px solid #22c55e;border-radius:8px;
     padding:8px 14px;margin-top:16px;margin-bottom:8px;font-size:15px;font-weight:800;color:#15803d">
  📦 不急件（出貨日 &gt; 10 工作天）
</div>""", unsafe_allow_html=True)
                for arr_date in sorted(normal_df["預計到料日"].unique()):
                    day_df = normal_df[normal_df["預計到料日"] == arr_date]
                    _date_header(arr_date, len(day_df))
                    _show_table(day_df)

        else:  # 急件 / 不急件 分開（左右欄）
            col_urg, col_norm = st.columns(2)
            urgent_df = has_date[has_date["_urgent"]].sort_values("預計到料日")
            normal_df = has_date[~has_date["_urgent"]].sort_values("預計到料日")

            with col_urg:
                st.markdown(f"""
<div style="background:#fff7ed;border:2px solid #f97316;border-radius:8px;
     padding:8px 14px;margin-bottom:10px;font-size:14px;font-weight:800;color:#c2410c">
  🚨 急件 &nbsp;<span style="font-weight:400;font-size:12px">共 {len(urgent_df)} 項</span>
</div>""", unsafe_allow_html=True)
                if urgent_df.empty:
                    st.success("無急件")
                else:
                    for arr_date in sorted(urgent_df["預計到料日"].unique()):
                        day_df = urgent_df[urgent_df["預計到料日"] == arr_date]
                        _date_header(arr_date, len(day_df))
                        _show_table(day_df)

            with col_norm:
                st.markdown(f"""
<div style="background:#f0fdf4;border:2px solid #22c55e;border-radius:8px;
     padding:8px 14px;margin-bottom:10px;font-size:14px;font-weight:800;color:#15803d">
  📦 不急件 &nbsp;<span style="font-weight:400;font-size:12px">共 {len(normal_df)} 項</span>
</div>""", unsafe_allow_html=True)
                if normal_df.empty:
                    st.info("無不急件")
                else:
                    for arr_date in sorted(normal_df["預計到料日"].unique()):
                        day_df = normal_df[normal_df["預計到料日"] == arr_date]
                        _date_header(arr_date, len(day_df))
                        _show_table(day_df)

        # 無日期的 IQC
        if not no_date.empty:
            st.markdown(f"""
<div style="background:#fefce8;border-left:5px solid #eab308;border-radius:6px;
     padding:8px 14px;margin-top:16px;margin-bottom:6px">
  <span style="font-size:15px;font-weight:800;color:#92400e">
    📦 IQC 驗收中（無明確到料日）
  </span>
  <span style="font-size:13px;color:#64748b;margin-left:10px">{len(no_date)} 項物料</span>
</div>""", unsafe_allow_html=True)
            _show_table(no_date)
