import streamlit as st
import pandas as pd
import numpy as np
import sys, os
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.shared import inject_css, render_header, render_sidebar

st.set_page_config(page_title="耗損率分析", page_icon="📉", layout="wide")
inject_css()
render_header(title="耗損率分析", subtitle="Loss Rate Analysis", badge="生管 PC")
render_sidebar()

st.markdown("---")

COST_FILE_DEFAULT = r"C:\Users\T26019\Desktop\材料金額.xlsx"

# ── 料號分類對照表（依料件編碼原則） ─────────────────────────────────────────
MATERIAL_CLASS = {
    # 材料-電子
    "2101": ("材料-電子", "電阻"),
    "2102": ("材料-電子", "電容"),
    "2103": ("材料-電子", "電感"),
    "2104": ("材料-電子", "電晶體"),
    "2105": ("材料-電子", "二極體"),
    "2106": ("材料-電子", "LED/LCD"),
    "2107": ("材料-電子", "振盪器"),
    "2108": ("材料-電子", "高頻電容/電感"),
    "2109": ("材料-電子", "FILTER"),
    "2111": ("材料-電子", "IC-Controller"),
    "2112": ("材料-電子", "IC-PHY"),
    "2113": ("材料-電子", "IC-CPU/MPU/RFMD"),
    "2114": ("材料-電子", "IC-MEMORY"),
    "2115": ("材料-電子", "TTL/CMOS"),
    "2116": ("材料-電子", "CLOCK/PLD"),
    "2117": ("材料-電子", "IC-LINEAR"),
    "2118": ("材料-電子", "電子零件"),
    "2131": ("材料-電子", "光纖收發器"),
    "2141": ("材料-電子", "連接器"),
    "2142": ("材料-電子", "保護元件"),
    "2143": ("材料-電子", "開關"),
    "2144": ("材料-電子", "Phone Jack"),
    "2145": ("材料-電子", "BUZZER"),
    "2151": ("材料-電子", "PCB"),
    "2161": ("材料-電子", "電源供應器"),
    "2162": ("材料-電子", "Transformer"),
    "2163": ("材料-電子", "Power Cord"),
    "2171": ("材料-電子", "線材類"),
    "2172": ("材料-電子", "ANTENNA"),
    "2173": ("材料-電子", "Hard Disk"),
    # 材料-軟體
    "1200": ("材料-軟體", "Royalty/權利金"),
    "1201": ("材料-軟體", "Driver/Utility"),
    "1202": ("材料-軟體", "CODE Pre-program"),
    "1203": ("材料-軟體", "Test Program"),
    "1206": ("材料-軟體", "F/W CODE+Test"),
    "1207": ("材料-軟體", "DOMAIN CODE"),
    "1208": ("材料-軟體", "F/W CODE"),
    "1209": ("材料-軟體", "On-line Programming"),
    # 材料-機構
    "1501": ("材料-機構", "BRACKET"),
    "1502": ("材料-機構", "PC CARD"),
    "1503": ("材料-機構", "散熱片"),
    "1504": ("材料-機構", "螺絲"),
    "1505": ("材料-機構", "風扇"),
    "1506": ("材料-機構", "扣件"),
    "1507": ("材料-機構", "腳墊"),
    "1508": ("材料-機構", "配件"),
    "1510": ("材料-機構", "CASE(ODM)"),
    "1511": ("材料-機構", "CASE(金屬)"),
    "1512": ("材料-機構", "CASE(塑膠)"),
    "1513": ("材料-機構", "CASE(防火塑膠)"),
    # 材料-包裝
    "1900": ("材料-包裝", "其他"),
    "1901": ("材料-包裝", "包材"),
    "1902": ("材料-包裝", "緩衝材(FOAM)"),
    "1903": ("材料-包裝", "袋子"),
    "1904": ("材料-包裝", "銘版(Nameplate)"),
    "1905": ("材料-包裝", "貼紙(Label)"),
    "1906": ("材料-包裝", "紙箱(CARTON)"),
    "1907": ("材料-包裝", "說明書(Manual)"),
    "1908": ("材料-包裝", "膠帶(Tape)"),
    "1911": ("材料-包裝", "光碟片(CD)"),
    # 半成品 / 成品
    "5145": ("半成品", "半成品"),
    "5100": ("成品", "成品"),
    "9000": ("成品", "成品"),
}

def classify(料號: str):
    prefix = str(料號).strip()[:4]
    result = MATERIAL_CLASS.get(prefix)
    if result:
        return result
    # 前2碼推斷大類
    p2 = prefix[:2]
    if p2 == "21":
        return ("材料-電子", f"電子({prefix})")
    if p2 == "15":
        return ("材料-機構", f"機構({prefix})")
    if p2 == "19":
        return ("材料-包裝", f"包裝({prefix})")
    if p2 == "12":
        return ("材料-軟體", f"軟體({prefix})")
    return ("其他", f"未知({prefix})")

# ── 載入單位成本 Lookup ───────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_cost_lookup(file_bytes: bytes | None, path: str) -> dict:
    try:
        src = BytesIO(file_bytes) if file_bytes else path
        if not file_bytes and not os.path.exists(path):
            return {}
        df = pd.read_excel(src, header=0)
        df_c = df[df.iloc[:, 0].notna() & df.iloc[:, 10].notna()].copy()
        lookup = df_c.groupby(df_c.columns[0])[df_c.columns[10]].first().to_dict()
        return {str(k).strip(): float(v) for k, v in lookup.items()}
    except Exception as e:
        st.warning(f"載入材料金額失敗：{e}")
        return {}


# ── 解析唐佑格式 (.xls, Sheet1, B=料號 col1, E=耗損數量 col4) ────────────────
@st.cache_data(show_spinner=False)
def load_tangyu(file_bytes: bytes, filename: str) -> pd.DataFrame:
    engine = "xlrd" if filename.lower().endswith(".xls") else "openpyxl"
    df = pd.read_excel(BytesIO(file_bytes), sheet_name="Sheet1", header=0, engine=engine)
    df = df.rename(columns={
        df.columns[1]: "料號",
        df.columns[3]: "品名",
        df.columns[4]: "耗損數量",
        df.columns[5]: "工單號",
        df.columns[6]: "使用機種",
    })
    df["料號"]     = df["料號"].astype(str).str.strip()
    df["品名"]     = df["品名"].astype(str).str.strip()
    df["耗損數量"] = pd.to_numeric(df["耗損數量"], errors="coerce").fillna(0).astype(int)
    df["工單號"]   = df["工單號"].astype(str).str.strip()
    df["使用機種"] = df["使用機種"].astype(str).str.strip()
    # 彙整（依料號+品名加總）
    agg = (
        df.groupby(["料號", "品名"], sort=False)
        .agg(耗損數量=("耗損數量", "sum"))
        .reset_index()
    )
    return agg, df[["料號", "品名", "耗損數量", "工單號", "使用機種"]]


# ── 解析國智格式 (.xlsx, 彙總 sheet index 2, A=料號 col0, B=數量 col1) ─────────
@st.cache_data(show_spinner=False)
def load_guozhi(file_bytes: bytes) -> pd.DataFrame:
    xl = pd.ExcelFile(BytesIO(file_bytes))
    df = pd.read_excel(xl, sheet_name=xl.sheet_names[2], header=0)
    df = df.rename(columns={df.columns[0]: "料號", df.columns[1]: "耗損數量"})
    df["料號"]     = df["料號"].astype(str).str.strip()
    df["耗損數量"] = pd.to_numeric(df["耗損數量"], errors="coerce").fillna(0).astype(int)
    df = df[df["料號"].notna() & (df["料號"] != "nan")].copy()
    return df[["料號", "耗損數量"]]


# ── 計算金額 + 分類 ───────────────────────────────────────────────────────────
def calc_amount(df: pd.DataFrame, lookup: dict) -> pd.DataFrame:
    df = df.copy()
    df["單位成本"] = df["料號"].map(lookup)
    df["耗損金額"] = (df["耗損數量"] * df["單位成本"]).round(4).fillna(0)
    df[["大類", "小類"]] = pd.DataFrame(
        df["料號"].apply(classify).tolist(), index=df.index
    )
    return df


# ── 顯示區塊 ──────────────────────────────────────────────────────────────────
def render_section(df: pd.DataFrame, detail_df: pd.DataFrame | None, label: str):
    total_skus  = len(df)
    total_qty   = int(df["耗損數量"].sum())
    total_amt   = df["耗損金額"].sum()
    matched_cnt = df["單位成本"].notna().sum()
    no_cost_cnt = df["單位成本"].isna().sum()

    k0, k1, k2, k3, k4 = st.columns(5)
    k0.metric("料號種類",   f"{total_skus:,}")
    k1.metric("耗損總數量", f"{total_qty:,}")
    k2.metric("耗損總金額", f"$ {total_amt:,.2f}")
    k3.metric("已計金額",   f"{matched_cnt} 種")
    k4.metric("未計金額",   f"{no_cost_cnt} 種",
              delta=f"需補單位成本" if no_cost_cnt else None,
              delta_color="inverse")

    # ── 分類圖表 ──────────────────────────────────────────────────────────────
    import altair as alt
    st.markdown("#### 📊 耗損金額分類分析")

    # 大類合計
    bigcat = (
        df.groupby("大類")
        .agg(耗損數量=("耗損數量", "sum"), 耗損金額=("耗損金額", "sum"))
        .reset_index()
        .sort_values("耗損金額", ascending=False)
    )
    bigcat["耗損金額_r"] = bigcat["耗損金額"].round(2)

    chart = (
        alt.Chart(bigcat)
        .mark_bar()
        .encode(
            x=alt.X("大類:N", sort="-y", title="大類", axis=alt.Axis(labelAngle=-20)),
            y=alt.Y("耗損金額_r:Q", title="耗損金額 (NT$)"),
            color=alt.Color("大類:N", legend=None),
            tooltip=["大類",
                     alt.Tooltip("耗損數量:Q", format=","),
                     alt.Tooltip("耗損金額_r:Q", title="耗損金額", format=",.2f")],
        )
        .properties(height=280)
    )
    st.altair_chart(chart, use_container_width=True)

    # 小類明細
    with st.expander("🗂️ 分類明細（大類 × 小類）"):
        cat_pivot = (
            df.groupby(["大類", "小類"])
            .agg(料號種類=("料號", "count"),
                 耗損數量=("耗損數量", "sum"),
                 耗損金額=("耗損金額", "sum"))
            .reset_index()
            .sort_values(["大類", "耗損金額"], ascending=[True, False])
        )
        cp_show = cat_pivot.copy()
        cp_show["耗損數量"] = cp_show["耗損數量"].apply(lambda x: f"{x:,}")
        cp_show["耗損金額"] = cp_show["耗損金額"].apply(lambda x: f"$ {x:,.2f}")
        st.dataframe(cp_show, use_container_width=True, hide_index=True)

    st.markdown("---")

    # 搜尋
    kw = st.text_input("🔍 料號 / 品名 搜尋", placeholder="輸入關鍵字...", key=f"kw_{label}")
    dff = df.copy()
    if kw:
        mask = dff["料號"].str.contains(kw, case=False, na=False)
        if "品名" in dff.columns:
            mask |= dff["品名"].str.contains(kw, case=False, na=False)
        dff = dff[mask]

    # 顯示表（含大類/小類）
    st.markdown(f"**耗損金額明細（{len(dff):,} 種）**")
    show = dff.copy()
    # 顯示欄位排序：大類/小類放前面
    base_cols = ["大類", "小類", "料號"]
    if "品名" in show.columns:
        base_cols.append("品名")
    base_cols += ["耗損數量", "單位成本", "耗損金額"]
    show = show[[c for c in base_cols if c in show.columns]]
    show["耗損數量"] = show["耗損數量"].apply(lambda x: f"{x:,}")
    show["單位成本"] = show["單位成本"].apply(lambda x: f"{x:,.4f}" if pd.notna(x) else "—")
    show["耗損金額"] = show["耗損金額"].apply(lambda x: f"$ {x:,.4f}" if x != 0 else "—")

    def row_style(row):
        if row["單位成本"] == "—":
            return ["background-color:#fdecea;color:#c62828"] * len(row)
        return [""] * len(row)

    st.dataframe(show.style.apply(row_style, axis=1),
                 use_container_width=True, height=440, hide_index=True)
    st.caption("🔴 紅底 = 找不到單位成本")

    filter_qty = int(dff["耗損數量"].astype(str).str.replace(",", "").astype(float).sum()) if kw else int(df["耗損數量"].sum())
    filter_amt = dff["耗損金額"].astype(str).str.replace("$ ", "").str.replace(",", "").astype(float).sum() if kw else df["耗損金額"].sum()
    st.markdown(f"**篩選合計：耗損數量 {int(dff['耗損數量'].astype(str).str.replace(',','').astype(float).sum()) if kw else total_qty:,} ／ 耗損金額 $ {dff['耗損金額'].astype(str).str.replace('$ ','').str.replace(',','').str.replace('—','0').astype(float).sum() if kw else total_amt:,.2f}**")

    # 工單明細（唐佑專用）
    if detail_df is not None:
        with st.expander("📋 查看工單原始明細（逐筆）"):
            raw = detail_df.copy()
            if kw:
                m2 = raw["料號"].str.contains(kw, case=False, na=False)
                if "品名" in raw.columns:
                    m2 |= raw["品名"].str.contains(kw, case=False, na=False)
                raw = raw[m2]
            raw["單位成本"] = raw["料號"].map(cost_lookup)
            raw["耗損金額"] = (raw["耗損數量"] * raw["單位成本"]).round(4).fillna(0)
            raw["耗損數量"] = raw["耗損數量"].apply(lambda x: f"{x:,}")
            raw["單位成本"] = raw["單位成本"].apply(lambda x: f"{x:,.4f}" if pd.notna(x) else "—")
            raw["耗損金額"] = raw["耗損金額"].apply(lambda x: f"$ {x:,.4f}" if x != 0 else "—")
            st.dataframe(raw, use_container_width=True, height=360, hide_index=True)

    return dff  # 回傳篩選後 df 供匯出使用


# ── Excel 匯出 ────────────────────────────────────────────────────────────────
def make_excel(datasets: list[tuple[str, pd.DataFrame]]) -> bytes:
    """datasets: list of (sheet_name, df with raw numeric values)"""
    wb = openpyxl.Workbook()
    HDR_F = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    HDR_T = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    TOT_F = PatternFill("solid", start_color="BBDEFB", end_color="BBDEFB")
    TOT_T = Font(bold=True, name="Arial", size=10)
    RED_F = PatternFill("solid", start_color="FDECEA", end_color="FDECEA")
    NRM_T = Font(name="Arial", size=10)
    THIN  = Side(style="thin", color="CCCCCC")
    BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CTR   = Alignment(horizontal="center", vertical="center")
    LFT   = Alignment(horizontal="left",   vertical="center")
    RGT   = Alignment(horizontal="right",  vertical="center")

    first = True
    for sheet_name, df in datasets:
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False

        cols = list(df.columns)
        widths = {"料號": 32, "品名": 38, "耗損數量": 12,
                  "工單號": 18, "使用機種": 40,
                  "單位成本": 14, "耗損金額": 16}

        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = HDR_F; c.font = HDR_T
            c.border = BDR; c.alignment = CTR
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 16)

        num_fmts = {"耗損數量": "#,##0", "單位成本": "#,##0.0000", "耗損金額": "#,##0.0000"}
        aligns   = {"料號": LFT, "品名": LFT, "工單號": CTR, "使用機種": LFT}

        for ri, row in enumerate(df.itertuples(index=False), 2):
            vals = list(row)
            no_cost = pd.isna(vals[cols.index("單位成本")]) if "單位成本" in cols else False
            fill = RED_F if no_cost else None
            for ci, (col, v) in enumerate(zip(cols, vals), 1):
                cell = ws.cell(row=ri, column=ci, value=None if (isinstance(v, float) and np.isnan(v)) else v)
                cell.font = NRM_T; cell.border = BDR
                cell.alignment = aligns.get(col, RGT)
                if col in num_fmts and v is not None:
                    cell.number_format = num_fmts[col]
                if fill:
                    cell.fill = fill

        # 合計列
        tr = len(df) + 2
        for ci in range(1, len(cols) + 1):
            c = ws.cell(row=tr, column=ci)
            c.fill = TOT_F; c.font = TOT_T; c.border = BDR
        ws.cell(row=tr, column=1, value="合計").alignment = CTR
        if "耗損數量" in cols:
            ci = cols.index("耗損數量") + 1
            ws.cell(row=tr, column=ci, value=int(df["耗損數量"].sum())).number_format = "#,##0"
            ws.cell(row=tr, column=ci).alignment = RGT
        if "耗損金額" in cols:
            ci = cols.index("耗損金額") + 1
            v = df["耗損金額"].fillna(0).sum()
            ws.cell(row=tr, column=ci, value=round(v, 4)).number_format = "#,##0.0000"
            ws.cell(row=tr, column=ci).alignment = RGT

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{tr-1}"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# 上傳區
# ════════════════════════════════════════════════════════════════════════════
uc1, uc2, uc3 = st.columns(3)
with uc1:
    tangyu_upload = st.file_uploader(
        "📂 唐佑：SMT/DIP耗損 (.xls)",
        type=["xls", "xlsx"], key="tangyu"
    )
with uc2:
    guozhi_upload = st.file_uploader(
        "📂 國智：工單耗損 (.xlsx)",
        type=["xlsx"], key="guozhi"
    )
with uc3:
    cost_upload = st.file_uploader(
        "📂 材料金額 (.xlsx)（選填）",
        type=["xlsx"], key="cost_file"
    )

if tangyu_upload is None and guozhi_upload is None:
    st.info("請上傳至少一份耗損明細檔案。")
    st.stop()

# 載入成本
cost_bytes = cost_upload.read() if cost_upload else None
cost_lookup = load_cost_lookup(cost_bytes, COST_FILE_DEFAULT)

# ════════════════════════════════════════════════════════════════════════════
# Tabs
# ════════════════════════════════════════════════════════════════════════════
tabs_labels = []
if tangyu_upload:
    tabs_labels.append("🔵 唐佑（SMT/DIP拋料）")
if guozhi_upload:
    tabs_labels.append("🟢 國智（工單耗損）")
if tangyu_upload and guozhi_upload:
    tabs_labels.append("📊 合計總覽")

tabs = st.tabs(tabs_labels)
tab_idx = 0

export_datasets = []  # [(sheet_name, df_raw)]

# ── 唐佑 Tab ─────────────────────────────────────────────────────────────────
if tangyu_upload:
    with tabs[tab_idx]:
        tab_idx += 1
        st.markdown("---")
        ty_bytes = tangyu_upload.read()
        ty_agg, ty_raw = load_tangyu(ty_bytes, tangyu_upload.name)
        ty_agg  = calc_amount(ty_agg,  cost_lookup)
        ty_raw2 = ty_raw.copy()
        ty_raw2["單位成本"] = ty_raw2["料號"].map(cost_lookup)
        ty_raw2["耗損金額"] = (ty_raw2["耗損數量"] * ty_raw2["單位成本"]).round(4).fillna(0)

        render_section(ty_agg, ty_raw, "tangyu")

        export_datasets.append(("唐佑_彙整", ty_agg))
        export_datasets.append(("唐佑_工單明細", ty_raw2))

# ── 國智 Tab ─────────────────────────────────────────────────────────────────
if guozhi_upload:
    with tabs[tab_idx]:
        tab_idx += 1
        st.markdown("---")
        gz_bytes = guozhi_upload.read()
        gz_df = load_guozhi(gz_bytes)
        gz_df = calc_amount(gz_df, cost_lookup)

        render_section(gz_df, None, "guozhi")

        export_datasets.append(("國智_彙整", gz_df))

# ── 合計總覽 Tab ──────────────────────────────────────────────────────────────
if tangyu_upload and guozhi_upload:
    with tabs[tab_idx]:
        st.markdown("---")
        st.markdown("### 兩份合計比較")

        rows = []
        if tangyu_upload:
            rows.append({
                "負責人": "唐佑（SMT/DIP）",
                "料號種類": len(ty_agg),
                "耗損總數量": int(ty_agg["耗損數量"].sum()),
                "耗損總金額": round(ty_agg["耗損金額"].sum(), 2),
                "未計金額種類": int(ty_agg["單位成本"].isna().sum()),
            })
        if guozhi_upload:
            rows.append({
                "負責人": "國智（工單耗損）",
                "料號種類": len(gz_df),
                "耗損總數量": int(gz_df["耗損數量"].sum()),
                "耗損總金額": round(gz_df["耗損金額"].sum(), 2),
                "未計金額種類": int(gz_df["單位成本"].isna().sum()),
            })

        total_qty_all = sum(r["耗損總數量"] for r in rows)
        total_amt_all = sum(r["耗損總金額"] for r in rows)
        rows.append({
            "負責人": "合計",
            "料號種類": sum(r["料號種類"] for r in rows[:-0 or len(rows)]),
            "耗損總數量": total_qty_all,
            "耗損總金額": total_amt_all,
            "未計金額種類": sum(r["未計金額種類"] for r in rows),
        })

        sum_df = pd.DataFrame(rows)
        sum_df["耗損總數量"] = sum_df["耗損總數量"].apply(lambda x: f"{x:,}")
        sum_df["耗損總金額"] = sum_df["耗損總金額"].apply(lambda x: f"$ {x:,.2f}")

        def hl_total(row):
            if row["負責人"] == "合計":
                return ["font-weight:bold;background-color:#e3f2fd"] * len(row)
            return [""] * len(row)

        st.dataframe(sum_df.style.apply(hl_total, axis=1),
                     use_container_width=True, hide_index=True)

        # KPI 大字
        ca, cb = st.columns(2)
        ca.metric("📦 耗損總數量（合計）", f"{total_qty_all:,}")
        cb.metric("💰 耗損總金額（合計）", f"$ {total_amt_all:,.2f}")

# ── 匯出按鈕 ─────────────────────────────────────────────────────────────────
if export_datasets:
    st.markdown("---")
    excel_bytes = make_excel(export_datasets)
    st.download_button(
        label="📥 匯出 Excel（所有明細）",
        data=excel_bytes,
        file_name="耗損金額明細.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
