import streamlit as st
import pandas as pd
import math
import io
import sys
import os
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.shared import ensure_calamine, inject_css, render_header, render_sidebar, render_sd_loader, render_nas_loader, source_is_csv
from utils.transfer_registry import (
    parse_registry, pending_map, status_map,
    NAS_REG_DIR, NAS_REG_PFX, PENDING_SHEETS,
)

_NAS_TRANSFER_DIR = "//192.168.2.34/MO_Storage/ORing MO/ORing-MO 工作/資材部/每日調撥與送燒ic(NEW)/生管互調料通知"
_NAS_TRANSFER_PFX = "加工廠互調料滙整表-"

ensure_calamine()

st.set_page_config(page_title="國智配料表", page_icon="🏭", layout="wide", initial_sidebar_state="expanded")
inject_css()
render_header(
    title="國智配料表",
    subtitle="Kuo Zhi Material Allocation List &nbsp;·&nbsp; ORing Industrial Networking",
    badge="Material Control · MC",
    show_logo=False,
)
render_sidebar()

TANG      = '唐佑代工倉'
KUO       = '修研/華盈/國智代工倉'
VALID_SRC = {'電子倉', '機構倉', '半成品倉', '成品倉'}  # 試算可用庫存只計這四個倉

# =========================
# Sidebar 設定
# =========================
with st.sidebar:
    st.divider()
    st.markdown("### ⚙️ 設定")

    h2o_file      = st.file_uploader("📂 上傳國智缺料表（W11）",          type=["xlsx", "xls", "csv"])
    sd_file       = render_sd_loader(key="guozhi")
    transfer_file = render_nas_loader(
        key="guozhi_transfer",
        nas_dir=_NAS_TRANSFER_DIR,
        prefix=_NAS_TRANSFER_PFX,
        label="📂 上傳加工廠互調料滙整表（選填覆蓋）",
        types=["xlsx", "xls", "xlsm"],
    )

    st.markdown("**🔁 新版互調料登記表**")
    st.caption("自動抓 `@新版互調料表` 內最新版，只讀待備料／待轉倉／待進料，**已完成不抓**")
    reg_file = render_nas_loader(
        key="guozhi_registry",
        nas_dir=NAS_REG_DIR,
        prefix=NAS_REG_PFX,
        label="📂 上傳 @新版-加工廠互調料彙整表（選填覆蓋）",
        types=["xlsx", "xlsm"],
    )
    if reg_file is not None:
        include_legacy = st.checkbox(
            "待調撥量併計舊版滙整表", value=False,
            help="新版登記表為主要來源。舊版滙整表若仍有未結案的調撥可勾選一併計入，"
                 "但同一筆料有可能被重複計算。",
        )
    else:
        include_legacy = True

    st.markdown("**📅 分析區間**")
    # 預設＝今天起 30 天（原本寫死 2026/5/1～5/31，日子一過就整張表都算不到缺料）
    _today = date.today()
    date_start = st.date_input("起始日", _today,                      format="YYYY/MM/DD")
    date_end   = st.date_input("結束日", _today + timedelta(days=30), format="YYYY/MM/DD")

    if date_end < date_start:
        st.error("⚠️ 結束日不可早於起始日！")

    days = (date_end - date_start).days + 1
    st.caption(f"共 **{days}** 天（{date_start} ～ {date_end}）")

    st.divider()
    st.info(
        "💡 **分析邏輯**\n\n"
        "從供需表中，找出料號在設定\n"
        "區間內，國智委外廠預計領用量：\n\n"
        f"- 🟢 **國智** → `{KUO}`"
    )

# =========================
# 主畫面
# =========================
if not h2o_file or not sd_file:
    st.info("👈 請在左側上傳「國智缺料表（W11）」及「供需表」開始分析")
    st.markdown("""
    <div style="background:#f0fdf4;border:1.5px dashed #86efac;border-radius:12px;padding:20px 24px;margin-top:16px;">
    <b style="color:#15803d;font-size:1rem;">📋 操作步驟</b>
    <ol style="color:#374151;margin-top:10px;line-height:2.2;">
      <li>國智提供的 <b>W11 缺料表</b>（Excel）直接上傳至左側<br>
          <span style="color:#6b7280;font-size:0.85rem;">↳ 欄位標題位於第 3 列；第 1~2 列為說明，系統自動略過</span></li>
      <li>ERP → 供需管理 → <b>供需表（分倉）</b> → 匯出 Excel，上傳至左側</li>
      <li>（選填）上傳<b>加工廠互調料彙整表</b>，追蹤已調撥進度</li>
      <li><b>新版互調料登記表</b>由 NAS <code>@新版互調料表</code> 自動載入：已登記的<b>待備料／待轉倉／待進料</b>（品號、數量）會併入待調撥量，<b>已完成不抓</b></li>
      <li>設定<b>分析區間</b>（起始日 ～ 結束日）</li>
      <li>系統自動對應 W11 缺料品號，從供需表找出國智倉預計結存</li>
    </ol>
    <br>
    <b style="color:#15803d;">🎯 分類邏輯</b>
    <table style="margin-top:8px;width:100%;border-collapse:collapse;font-size:0.88rem;">
      <tr style="background:#dcfce7;"><td style="padding:5px 10px;">🟢 充足</td><td style="padding:5px 10px;">國智代工倉在設定區間內預計結存 ≥ 0</td></tr>
      <tr><td style="padding:5px 10px;">🔴 缺料</td><td style="padding:5px 10px;">預計結存 &lt; 0，需從廠內倉調撥至國智代工倉</td></tr>
      <tr style="background:#dcfce7;"><td style="padding:5px 10px;">🟡 已調撥</td><td style="padding:5px 10px;">互調料彙整表中有記錄，調撥後結存轉為充足</td></tr>
      <tr><td style="padding:5px 10px;">🔁 已登記</td><td style="padding:5px 10px;">新版互調料表已登記 <code>待備料 / 待轉倉 / 待進料</code>，不必重開調撥單（<code>已完成</code>不列入）</td></tr>
      <tr><td style="padding:5px 10px;">⚪ 未找到</td><td style="padding:5px 10px;">W11 料號在供需表中無記錄，需人工確認</td></tr>
    </table>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

with st.spinner("分析中，請稍候..."):
    # 讀國智缺料表（第3行為欄位標題，前2行為說明列）
    try:
        if h2o_file.name.endswith('.csv'):
            gz = None
            for enc in ['utf-8-sig', 'cp950', 'big5']:
                try:
                    gz = pd.read_csv(h2o_file, header=0, encoding=enc)
                    break
                except Exception:
                    h2o_file.seek(0)
            if gz is None:
                st.error("國智缺料表 CSV 無法讀取，請確認編碼。")
                st.stop()
        else:
            # W11 缺料表：第2張工作表，第1列即為標題（header=0），A欄='品號'
            try:
                gz = pd.read_excel(h2o_file, sheet_name=1, header=0, engine='calamine')
            except Exception:
                h2o_file.seek(0)
                gz = pd.read_excel(h2o_file, sheet_name=1, header=0, engine='openpyxl')
    except Exception as e:
        st.error(f"國智缺料表讀取失敗：{e}")
        st.stop()

    # 讀供需表
    try:
        if source_is_csv(sd_file):
            for enc in ['utf-8-sig', 'cp950', 'big5']:
                try:
                    sd = pd.read_csv(sd_file, header=0, encoding=enc)
                    break
                except Exception:
                    if hasattr(sd_file, 'seek'): sd_file.seek(0)
        else:
            sd = pd.read_excel(sd_file, sheet_name=0, header=0)
        sd['日期'] = pd.to_datetime(sd['日期'], errors='coerce')
    except Exception as e:
        st.error(f"供需表讀取失敗：{e}")
        st.stop()

    # 欄位檢查
    if '品號' not in gz.columns:
        st.error(f"國智缺料表找不到「品號」欄位，偵測到的欄位：{gz.columns.tolist()}")
        st.stop()
    for col in ['品號', '庫別名稱', '日期', '異動別', '異動數量']:
        if col not in sd.columns:
            st.error(f"供需表找不到「{col}」欄位，請確認檔案格式。")
            st.stop()

    start = pd.Timestamp(date_start)
    end   = pd.Timestamp(date_end)

    # ── 讀取加工廠互調料滙整表（選填）──
    # H欄(index 7)=料號、J欄(index 9)=國智代工倉待調撥量、K欄(index 10)=唐佑代工倉待調撥量
    kuo_pending_map = {}
    if transfer_file is not None:
        try:
            tf_raw = pd.read_excel(transfer_file, sheet_name=0, header=None, engine='openpyxl')
            tf = tf_raw.iloc[1:].copy()
            tf_h = tf[7].astype(str).str.strip()
            tf_j = pd.to_numeric(tf[9], errors='coerce').fillna(0)
            tf_df = pd.DataFrame({'料號': tf_h, 'J': tf_j})
            tf_df = tf_df[tf_df['料號'].notna() & (tf_df['料號'] != '') & (tf_df['料號'] != 'nan')]
            kuo_pending_map = tf_df.groupby('料號')['J'].sum().to_dict()
        except Exception as e:
            st.warning(f"加工廠互調料滙整表讀取失敗（略過）：{e}")

    # ── 讀取新版互調料登記表（只抓未完成的待備料／待轉倉／待進料）──
    @st.cache_data(show_spinner=False, ttl=600)
    def _parse_registry_cached(cache_key: str, data: bytes):
        return parse_registry(io.BytesIO(data))

    def _load_registry(src):
        if src is None:
            return None, []
        try:
            if isinstance(src, str):
                key = f"{src}|{os.path.getmtime(src)}"
                with open(src, 'rb') as f:
                    data = f.read()
            else:
                data = src.getvalue()
                key = f"upload|{getattr(src, 'name', '')}|{len(data)}"
        except Exception as e:
            return None, [f"檔案讀取失敗：{e}"]
        return _parse_registry_cached(key, data)

    reg_df, reg_warns = _load_registry(reg_file)
    for _w in reg_warns:
        st.warning(f"新版互調料登記表：{_w}")

    kuo_reg_map = pending_map(reg_df, '國智')
    kuo_reg_txt = status_map(reg_df, '國智')
    has_registry = bool(kuo_reg_map)

    # 待調撥量 = 新版登記表未完成量 ＋（可選）舊版滙整表
    if has_registry and not include_legacy:
        kuo_pending_map = {}
    for _k, _v in kuo_reg_map.items():
        kuo_pending_map[_k] = (kuo_pending_map.get(_k, 0) or 0) + _v

    has_transfer = bool(kuo_pending_map)

    spq_map = {}
    if 'SPQ' in sd.columns:
        spq_map = (sd[sd['SPQ'].notna()][['品號','SPQ']]
                   .drop_duplicates('品號')
                   .set_index('品號')['SPQ']
                   .to_dict())

    def apply_spq(qty, spq):
        s = int(spq) if spq and spq > 0 else 1
        if qty <= 0: return 0
        return math.ceil(qty / s) * s

    def end_deficit(df_sd, pno, wh_name):
        sub = df_sd[(df_sd['品號']==pno) & (df_sd['庫別名稱']==wh_name) &
                    (df_sd['日期'] <= end) & df_sd['預計結存'].notna()]
        if sub.empty: return 0
        last_bal = sub.sort_values('日期').iloc[-1]['預計結存']
        return max(0, -last_bal)

    def get_avail(df_sd, pno, wh_code, excl):
        w = df_sd[(df_sd['品號']==pno) & (df_sd['庫別']==wh_code)]
        if w.empty: return 0
        wh_name = w['庫別名稱'].dropna().iloc[0] if w['庫別名稱'].dropna().shape[0]>0 else ''
        if wh_name in excl: return 0
        dated = w[w['日期'].notna() & w['預計結存'].notna()]
        in_range = dated[dated['日期'] <= end]
        if not in_range.empty:
            last_bal = in_range.sort_values('日期').iloc[-1]['預計結存']
            # 預計進貨 & 預計生產：尚未實際入庫/完工，不算現有可用庫存，須扣除
            planned_in = dated[
                (dated['日期'] <= end) &
                (dated['異動別'].isin(['預計進貨', '預計生產']))
            ]['異動數量'].sum()
            return max(0, last_bal - planned_in)
        init_rows = w[w['日期'].isna()]
        if not init_rows.empty:
            qty = init_rows['異動數量'].dropna()
            if not qty.empty:
                return max(0, qty.iloc[0])
            qty = init_rows['預計結存'].dropna()
            if not qty.empty:
                return max(0, qty.iloc[0])
        return 0

    def source_wh(df_sd, pno, exclude_names, need_qty):
        excl = set(exclude_names)
        part_sd = df_sd[df_sd['品號']==pno]
        result = []
        remaining = need_qty

        # 有日期資料的列：建立 代碼→名稱 對照
        dated_part = part_sd[part_sd['日期'].notna() & part_sd['庫別名稱'].notna()]
        code_name = (dated_part[['庫別','庫別名稱']]
                     .drop_duplicates('庫別')
                     .set_index('庫別')['庫別名稱']
                     .to_dict())
        code_name = {c: n for c, n in code_name.items() if len(str(c)) <= 12}
        names_with_dated = set(code_name.values())
        dated_codes      = set(code_name.keys())

        # 只有初始列的倉（無日期、無倉名、不在有交易的倉之中）
        init_rows = part_sd[part_sd['日期'].isna() & part_sd['庫別名稱'].isna()]
        init_only = {}
        for wh_k in init_rows['庫別'].dropna().unique():
            ws = str(wh_k)
            if ws in dated_codes or ws in names_with_dated: continue
            if len(ws) > 12: continue
            qty = init_rows[init_rows['庫別']==wh_k]['異動數量'].dropna()
            if not qty.empty and qty.iloc[0] > 0:
                init_only[ws] = qty.iloc[0]

        # Step 1: 電子倉優先（有交易→代碼查；無交易→初始列直讀）
        e_code = next((c for c, n in code_name.items() if n == '電子倉'), None)
        e_avail = get_avail(df_sd, pno, e_code, excl) if e_code else init_only.get('電子倉', 0)
        if e_avail > 0:
            result.append(f"電子倉（{int(e_avail):,}）")
            remaining -= min(e_avail, remaining)

        if remaining <= 0:
            return '、'.join(result)

        # Step 2: 有交易的其他倉
        for wh_code, wh_name in code_name.items():
            if e_code and wh_code == e_code: continue
            if wh_name in excl: continue
            avail = get_avail(df_sd, pno, wh_code, excl)
            if avail > 0:
                result.append(f"{wh_name}（{int(avail):,}）")
                remaining -= avail
                if remaining <= 0: break

        # Step 3: 只有初始列的其他倉
        for wh_name, avail in init_only.items():
            if wh_name == '電子倉': continue
            if wh_name in excl: continue
            if avail > 0:
                result.append(f"{wh_name}（{int(avail):,}）")
                remaining -= avail
                if remaining <= 0: break

        return '、'.join(result) if result else ''

    def src_avail_excl(pno, excl_names):
        """排除指定倉名後，所有可用來源倉的可用量加總（與 source_wh 相同識別邏輯）"""
        excl = set(excl_names)
        part_sd = sd[sd['品號']==pno]
        total = 0

        dated_part = part_sd[part_sd['日期'].notna() & part_sd['庫別名稱'].notna()]
        code_name = (dated_part[['庫別','庫別名稱']]
                     .drop_duplicates('庫別')
                     .set_index('庫別')['庫別名稱']
                     .to_dict())
        code_name = {c: n for c, n in code_name.items() if len(str(c)) <= 12}
        names_with_dated = set(code_name.values())
        dated_codes      = set(code_name.keys())

        for wh_code, wh_name in code_name.items():
            if wh_name in excl: continue
            total += get_avail(sd, pno, wh_code, excl)

        init_rows = part_sd[part_sd['日期'].isna() & part_sd['庫別名稱'].isna()]
        for wh_k in init_rows['庫別'].dropna().unique():
            ws = str(wh_k)
            if ws in dated_codes or ws in names_with_dated: continue
            if ws in excl: continue
            if len(ws) > 12: continue
            qty = init_rows[init_rows['庫別']==wh_k]['異動數量'].dropna()
            if not qty.empty and float(qty.iloc[0]) > 0:
                total += float(qty.iloc[0])

        return int(total)

    def first_deficit_date(pno, wh_name):
        sub = sd[(sd['品號']==pno) & (sd['庫別名稱']==wh_name) &
                 (sd['日期'] >= start) & (sd['日期'] <= end) &
                 sd['預計結存'].notna() & (sd['預計結存'] < 0)]
        if sub.empty: return None
        return sub.sort_values('日期').iloc[0]['日期']

    def get_incoming(pno):
        """取得所有預計進貨（不限分析區間），每筆格式：MM/DD(數量)"""
        sub = sd[
            (sd['品號'] == pno) &
            (sd['異動別'] == '預計進貨') &
            (sd['日期'].notna())
        ].sort_values('日期')
        if sub.empty:
            return None
        return '、'.join(
            f"{row['日期'].strftime('%m/%d')}({int(row['異動數量'])})"
            for _, row in sub.iterrows()
        )

    def avail_4wh(pno, excl_names):
        """只計算 VALID_SRC（電子倉/機構倉/半成品倉/成品倉）的可用量。"""
        excl = set(excl_names)
        part_sd = sd[sd['品號'] == pno]
        total = 0.0

        dated_part = part_sd[part_sd['日期'].notna() & part_sd['庫別名稱'].notna()]
        code_name = (dated_part[['庫別', '庫別名稱']]
                     .drop_duplicates('庫別')
                     .set_index('庫別')['庫別名稱']
                     .to_dict())
        code_name = {c: n for c, n in code_name.items() if len(str(c)) <= 12}
        names_with_dated = set(code_name.values())
        dated_codes      = set(code_name.keys())

        for wh_code, wh_name in code_name.items():
            if wh_name not in VALID_SRC: continue
            if wh_name in excl: continue
            total += get_avail(sd, pno, wh_code, excl)

        init_rows = part_sd[part_sd['日期'].isna() & part_sd['庫別名稱'].isna()]
        for wh_k in init_rows['庫別'].dropna().unique():
            ws = str(wh_k)
            if ws in dated_codes or ws in names_with_dated: continue
            if ws not in VALID_SRC: continue
            if ws in excl: continue
            if len(ws) > 12: continue
            qty = init_rows[init_rows['庫別'] == wh_k]['異動數量'].dropna()
            if not qty.empty and float(qty.iloc[0]) > 0:
                total += float(qty.iloc[0])

        return int(total)

    parts = gz['品號'].dropna().unique()

    rows = []
    for pno in parts:
        gz_row  = gz[gz['品號']==pno].iloc[0]
        spq     = spq_map.get(pno, 1)
        spq_int = int(spq) if spq and spq > 0 else 1

        k_deficit = end_deficit(sd, pno, KUO)
        k_qty     = apply_spq(k_deficit, spq)   # 顯示用（SPQ進位）

        if k_deficit <= 0:
            continue  # 只顯示國智有缺料的料號

        # ── 待調撥量（優先計算，用於推算淨需求）──
        pno_str   = str(pno).strip()
        k_pending = int(kuo_pending_map.get(pno_str, 0) or 0)

        # 淨需求 = 原始缺料量 - 已調撥量（不套 SPQ）
        k_net = max(0, k_deficit - k_pending)

        # ── 可調撥庫存：只計算四個倉，扣除已待調撥量後的剩餘可用量 ──
        avail_4w     = avail_4wh(pno, {KUO}) if k_net > 0 else 0
        net_avail_4w = max(0, avail_4w - k_pending)

        # ── 可調撥來源倉（顯示用，仍顯示所有倉的現有庫存）──
        src = source_wh(sd, pno, set(), k_qty)

        # ── 配料分配（三段邏輯）──
        k_spq_net  = apply_spq(k_net, spq)
        k_alloc    = k_net
        short_note = ''
        shortage   = (k_net > 0 and net_avail_4w < k_net)

        if shortage:
            # 庫存連淨需求都不夠：給現有，不套 SPQ
            k_alloc = min(k_net, net_avail_4w)
            _pend_note = f"已待調撥 {k_pending:,}，" if k_pending > 0 else ""
            short_note = (
                f"⚠️ 庫存不足（原缺 {int(k_deficit):,}，{_pend_note}"
                f"尚缺 {int(k_net):,}，四倉剩餘可調 {int(net_avail_4w):,}，差 {int(k_net - k_alloc):,}）\n"
                f"► 國智代工倉 → 僅可配 {int(k_alloc):,}，尚缺 {int(k_net - k_alloc):,}"
            )
        elif net_avail_4w >= k_spq_net:
            # 剩餘庫存足夠 SPQ 進位量：給整包 SPQ
            k_alloc = k_spq_net
        # else: 只夠淨需求但不夠 SPQ 進位 → 維持 k_net

        # 缺料量顯示
        if shortage:
            _gap = int(k_net - k_alloc)
            if k_pending > 0:
                k_qty_str = (f"⚠️ 原缺{int(k_deficit):,}｜已調{k_pending:,}｜"
                             f"尚缺{int(k_net):,}｜可調{int(k_alloc):,}｜差{_gap:,}")
            else:
                k_qty_str = f"⚠️ 尚缺{int(k_net):,}｜可調{int(k_alloc):,}｜差{_gap:,}"
        else:
            d = int(k_deficit)
            k_qty_str = f"{k_qty:,}  (原缺 {d:,})" if k_qty != d else k_qty

        # 實際應調撥量：待調撥 >= 原缺 → 0；否則給分配量
        k_actual = (0 if k_pending >= k_deficit else k_alloc) if k_deficit > 0 else None

        # 客戶料號 & 預計進貨日（不限分析區間）
        cust_pn  = str(gz_row.get('客戶料號', '') or '') if pd.notna(gz_row.get('客戶料號', '')) else ''
        incoming = get_incoming(pno)

        rows.append({
            '品號':               pno,
            'SPQ':                spq_int,
            '國智代工倉 缺料量':   k_qty_str,
            '國智代工倉 待調撥量':    k_pending if (has_transfer and k_deficit > 0) else None,
            '國智代工倉 實際應調撥量': k_actual  if has_transfer else None,
            '國智 登記明細':          kuo_reg_txt.get(pno_str, '') if has_registry else None,
            '預計進貨日':          incoming,
            '_國智qty':           k_qty,
            '_shortage':          shortage,
            '可調撥來源倉（倉代碼/可用量）': src,
            '⚠️ 配料說明':        short_note,
            '客戶料號':            cust_pn,
        })

    df_out = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['品號','SPQ','國智代工倉 缺料量','國智代工倉 待調撥量','國智代工倉 實際應調撥量',
                 '國智 登記明細','_國智qty','_shortage',
                 '可調撥來源倉（倉代碼/可用量）','⚠️ 配料說明','客戶料號']
    )

# =========================
# 統計卡片
# =========================
short_warn = df_out[df_out['_shortage'] == True] if len(df_out) else pd.DataFrame()
_mcols = st.columns(5 if has_registry else 4)
col1, col2, col3, col4 = _mcols[:4]
col1.metric("缺料表料號總數",    f"{len(gz['品號'].dropna().unique())} 個")
col2.metric("國智有缺料料號",   f"{len(df_out)} 個")
col3.metric("國智總缺料量",     f"{int(df_out['_國智qty'].sum()):,}" if len(df_out) else "0")
col4.metric("⚠️ 庫存不足料號",  f"{len(short_warn)} 個",
            delta=None if len(short_warn)==0 else "需確認配料",
            delta_color="inverse")
if has_registry:
    _reg_hit = sum(1 for _p in parts if str(_p).strip() in kuo_reg_map)
    _mcols[4].metric("🔁 已登記待調撥", f"{_reg_hit} 個",
                     delta=f"登記表未完成 {len(reg_df):,} 筆", delta_color="off")

st.divider()

# =========================
# 資料表
# =========================
st.markdown(f"#### 🏭 國智配料表（區間末結存）　{date_start} ～ {date_end}")

# =========================
# 新版互調料登記表：未完成登記明細（已完成不抓）
# =========================
def render_registry_section():
    """畫出未完成登記明細，並回傳畫面上那張表（供匯出用）；沒資料回 None"""
    if reg_df is None or reg_df.empty:
        return None

    st.divider()
    st.markdown("#### 🔁 新版互調料表　未完成登記明細")
    _cnt = reg_df['狀態'].value_counts()
    st.caption(
        "　｜　".join(f"**{_s}** {int(_cnt.get(_s, 0)):,} 筆" for _s in PENDING_SHEETS)
        + f"　｜　⏰ 逾期 {int(reg_df['逾期'].sum()):,} 筆"
        + f"　｜　🔥 急料 {int(reg_df['急料'].sum()):,} 筆"
        + "　（「已完成」工作表不列入）"
    )

    _f1, _f2, _f3 = st.columns([2.2, 2.2, 1.6])
    with _f1:
        _sts = st.multiselect("狀態", PENDING_SHEETS, default=PENDING_SHEETS, key="gz_reg_sts")
    with _f2:
        _facs = sorted(reg_df['廠別'].unique().tolist())
        _fdef = ['國智'] if '國智' in _facs else _facs
        _fac  = st.multiselect("調入廠別", _facs, default=_fdef, key="gz_reg_fac")
    with _f3:
        _only_pn   = st.checkbox("只看本表料號", value=True,  key="gz_reg_only_pn")
        _only_late = st.checkbox("只看逾期",     value=False, key="gz_reg_only_late")

    rv = reg_df[reg_df['狀態'].isin(_sts) & reg_df['廠別'].isin(_fac)]
    if _only_pn:
        rv = rv[rv['品號'].isin({str(x).strip() for x in parts})]
    if _only_late:
        rv = rv[rv['逾期']]

    if rv.empty:
        st.info("目前條件下沒有未完成的登記。")
        return None

    reg_view = rv.copy()
    for _c in ['開單日', '通知日', '填表日期', '預計進貨日', '最終完成日']:
        reg_view[_c] = reg_view[_c].dt.date
    reg_view['數量']     = reg_view['數量'].astype('Int64')
    reg_view['逾期天數'] = reg_view['逾期天數'].apply(
        lambda v: f"⏰ 逾期 {int(v)} 天" if pd.notna(v) else '')
    reg_view['急料']     = reg_view['急料'].map({True: '🔥 急料', False: ''})
    reg_view = reg_view[['狀態','廠別','品號','數量','轉出倉別','單號及領料單號',
                         '開單日','通知日','預計進貨日','最終完成日',
                         '填表人','逾期天數','急料','備註']]

    _own_pnos = {str(x).strip() for x in parts}

    def _reg_row_style(row):
        if row['逾期天數']:
            return ['background-color:#fce7f3; color:#9d174d; font-weight:600;'] * len(row)
        if row['品號'] in _own_pnos:
            return ['background-color:#eff6ff; color:#1e3a8a;'] * len(row)
        return [''] * len(row)

    st.dataframe(
        reg_view.style.apply(_reg_row_style, axis=1),
        use_container_width=True,
        height=400,
        hide_index=True,
    )
    st.caption(
        f"顯示 {len(reg_view):,} 筆　·　"
        "粉紅＝逾期（待備料 最終完成日 >3 天、待轉倉 通知日 >7 天、待進料 預計進貨日 >2 天）"
        "　·　淺藍＝本表料號"
    )
    return reg_view


if df_out.empty:
    st.success("✅ 區間內國智代工倉無缺料！")
    render_registry_section()
else:
    # 庫存不足警告橫幅
    if len(short_warn) > 0:
        pno_list = '、'.join(short_warn['品號'].tolist())
        st.warning(
            f"**⚠️ 以下 {len(short_warn)} 個料號庫存不足，請確認配料：**\n\n{pno_list}",
            icon="⚠️",
        )

    if has_transfer:
        display_cols = ['品號','SPQ','國智代工倉 缺料量','國智代工倉 待調撥量']
        if has_registry:
            display_cols += ['國智 登記明細']
        display_cols += ['國智代工倉 實際應調撥量','預計進貨日',
                         '可調撥來源倉（倉代碼/可用量）','⚠️ 配料說明','客戶料號']
    else:
        display_cols = ['品號','SPQ','國智代工倉 缺料量',
                        '預計進貨日',
                        '可調撥來源倉（倉代碼/可用量）','⚠️ 配料說明','客戶料號']
    df_display = df_out[display_cols].copy()
    # 待調撥量／實際應調撥量欄內混有 None 會被 pandas 轉成 float，顯示成 6000.0 → 去小數
    for _c in ['國智代工倉 待調撥量', '國智代工倉 實際應調撥量']:
        if _c in df_display.columns:
            df_display[_c] = df_display[_c].apply(
                lambda v: f"{int(float(v)):,}" if isinstance(v, (int, float)) else v)

    _shortage_pnos = set(short_warn['品號'].tolist()) if len(short_warn) else set()

    def _row_style(row):
        pno = row.get('品號', '')
        if pno in _shortage_pnos:
            return ['background-color: #fef2f2; color: #991b1b; font-weight:600;'] * len(row)
        src_val = str(row.get('可調撥來源倉（倉代碼/可用量）', ''))
        if '、' in src_val:
            return ['background-color: #fff7ed; color: #9a3412; font-weight:600;'] * len(row)
        return [''] * len(row)

    st.dataframe(
        df_display.style.apply(_row_style, axis=1),
        use_container_width=True,
        height=520,
        hide_index=True,
        column_config={
            '⚠️ 配料說明': st.column_config.TextColumn(width='large'),
        }
    )

    reg_view = render_registry_section()

    # =========================
    # 匯出 Excel
    # =========================
    # 各欄底色/字色：(底色, 字色, 粗體)；用欄位鍵比對，插欄不會跑位
    _XL_KEY_STYLE = {
        '國智代工倉 缺料量':       ('FFD9E8FF', 'FF1E3A8A', True),
        '國智代工倉 待調撥量':     ('FFB8D4EE', 'FF1E3A8A', True),
        '國智 登記明細':           ('FFE4EEF9', 'FF1E3A8A', False),
        '國智代工倉 實際應調撥量': ('FFA0C4E8', 'FF0F2460', True),
        '預計進貨日':              ('FFE8F4FD', None, False),
        '可調撥來源倉（倉代碼/可用量）': ('FFFFF0CC', None, False),
    }

    def build_excel(df, start, end, with_transfer=False, with_registry=False, reg_rows=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '國智配料表'
        thin   = Side(style='thin', color='FFCCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # specs：(表頭, 表頭底色, 資料欄鍵, 欄寬, 對齊)
        specs = [
            ('品號',                'FFD9E8FF', '品號',               28, 'left'),
            ('SPQ',                 'FFF2F2F2', 'SPQ',                 8, 'center'),
            ('國智代工倉\n缺料量',   'FFD9E8FF', '國智代工倉 缺料量',   16, 'center'),
        ]
        if with_transfer:
            specs.append(('國智代工倉\n待調撥量', 'FFB8D4EE', '國智代工倉 待調撥量', 14, 'center'))
            if with_registry:
                specs.append(('國智\n登記明細', 'FFE4EEF9', '國智 登記明細', 22, 'left'))
            specs.append(('國智代工倉\n實際應調撥量', 'FFA0C4E8', '國智代工倉 實際應調撥量', 16, 'center'))
        specs += [
            ('預計進貨日',                     'FFE8F4FD', '預計進貨日',                    24, 'left'),
            ('可調撥來源倉\n（倉代碼/可用量）', 'FFF5E6FF', '可調撥來源倉（倉代碼/可用量）', 36, 'left'),
            ('配料說明\n（庫存不足時）',        'FFFCE4D6', '⚠️ 配料說明',                   40, 'left'),
            ('客戶料號',                       'FFF2F2F2', '客戶料號',                      28, 'left'),
        ]

        headers    = [x[0] for x in specs]
        hdr_color  = [x[1] for x in specs]
        col_order  = [x[2] for x in specs]
        col_widths = [x[3] for x in specs]
        left_cols  = {i for i, x in enumerate(specs, 1) if x[4] == 'left'}
        total_cols = len(specs)

        merge_end = get_column_letter(total_cols)
        ws.merge_cells(f'A1:{merge_end}1')
        c = ws['A1']
        c.value = f'國智配料表　{start.strftime("%Y/%m/%d")} ～ {end.strftime("%Y/%m/%d")}'
        c.font  = Font(name='Arial', bold=True, size=12, color='FFFFFFFF')
        c.fill  = PatternFill('solid', start_color='FF1E3A8A')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24

        for i, (h, hc) in enumerate(zip(headers, hdr_color), 1):
            cell = ws.cell(row=2, column=i, value=h)
            cell.font  = Font(name='Arial', bold=True, size=9)
            cell.fill  = PatternFill('solid', start_color=hc)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 32

        for r_i, row_dict in enumerate(df.to_dict('records'), 3):
            is_short = bool(row_dict.get('_shortage', False))
            for c_i, key in enumerate(col_order, 1):
                val  = row_dict.get(key)
                cell = ws.cell(row=r_i, column=c_i, value=val)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='left' if c_i in left_cols else 'center',
                    vertical='center',
                    wrap_text=(key == '⚠️ 配料說明' or key.endswith('登記明細')),
                )
                cell.font = Font(name='Arial', size=9, bold=is_short,
                                 color='FF991B1B' if is_short else 'FF000000')

                style  = _XL_KEY_STYLE.get(key)
                # 待調撥量／實際應調撥量：空白也上色，且缺料列也維持自己的藍色系
                always = key.endswith('待調撥量') or key.endswith('實際應調撥量')
                if style and (val or always) and (always or not is_short):
                    fc, tc, bold = style
                    cell.fill = PatternFill('solid', start_color=fc)
                    if tc:
                        cell.font = Font(name='Arial', size=9, bold=bold, color=tc)
                elif is_short:
                    cell.fill = PatternFill('solid', start_color='FFFCE4EC')
                elif key == '⚠️ 配料說明' and val:
                    cell.fill = PatternFill('solid', start_color='FFFDE8D0')
                    cell.font = Font(name='Arial', size=8, bold=True, color='FFC0392B')
                    ws.row_dimensions[r_i].height = 52

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A3'

        # ── 第 2 張表：新版互調料表未完成登記明細（已完成不列入）──
        if reg_rows is not None and not reg_rows.empty:
            ws2   = wb.create_sheet('未完成互調料登記')
            rcols = list(reg_rows.columns)
            rw    = {'狀態': 10, '廠別': 10, '品號': 32, '數量': 12, '轉出倉別': 12,
                     '單號及領料單號': 20, '開單日': 12, '通知日': 12, '預計進貨日': 12,
                     '最終完成日': 12, '填表人': 10, '逾期天數': 14, '急料': 10, '備註': 22}
            end_col = get_column_letter(len(rcols))
            ws2.merge_cells(f'A1:{end_col}1')
            t = ws2['A1']
            t.value = '新版互調料表　未完成登記明細（待備料／待轉倉／待進料，已完成不列入）'
            t.font  = Font(name='Arial', bold=True, size=12, color='FFFFFFFF')
            t.fill  = PatternFill('solid', start_color='FF1E3A8A')
            t.alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[1].height = 24

            for i, h in enumerate(rcols, 1):
                cell = ws2.cell(row=2, column=i, value=h)
                cell.font  = Font(name='Arial', bold=True, size=9)
                cell.fill  = PatternFill('solid', start_color='FFD9E8FF')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            ws2.row_dimensions[2].height = 22

            pink = PatternFill('solid', start_color='FFFCE7F3')
            for r_i, rec in enumerate(reg_rows.to_dict('records'), 3):
                late = bool(str(rec.get('逾期天數') or '').strip())
                for c_i, key in enumerate(rcols, 1):
                    v = rec.get(key)
                    if v is not None and not isinstance(v, str) and pd.isna(v):
                        v = None
                    cell = ws2.cell(row=r_i, column=c_i, value=v)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal='left' if key in ('品號', '單號及領料單號', '備註') else 'center',
                        vertical='center')
                    if late:
                        cell.fill = pink
                        cell.font = Font(name='Arial', size=9, bold=True, color='FF9D174D')
                    else:
                        cell.font = Font(name='Arial', size=9)

            for i, h in enumerate(rcols, 1):
                ws2.column_dimensions[get_column_letter(i)].width = rw.get(h, 14)
            ws2.freeze_panes = 'A3'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    buf = build_excel(df_out, start, end, with_transfer=has_transfer,
                      with_registry=has_registry, reg_rows=reg_view)
    st.download_button(
        label="⬇️ 匯出國智配料表（Excel）",
        data=buf,
        file_name=f"國智配料表_{date_start.strftime('%Y%m%d')}_{date_end.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
