import streamlit as st
import pandas as pd
import math
import io
import sys
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.shared import ensure_calamine, inject_css, render_header, render_sidebar

ensure_calamine()

st.set_page_config(page_title="廠內配料表", page_icon="🏭", layout="wide", initial_sidebar_state="expanded")
inject_css()
render_header(
    title="廠內配料表",
    subtitle="Factory Internal Material Allocation &nbsp;·&nbsp; ORing Industrial Networking",
    badge="Material Control · MC",
    show_logo=False,
)
render_sidebar()

VALID_SRC = {'電子倉', '機構倉', '半成品倉', '成品倉', '生產加工倉', '包材倉'}

# 廠內排程.xlsx 欄位位置（0-indexed，第1列為欄名，跳過）
#   A(0)=料號  B(1)=品名  F(5)=工單單號  H(7)=需求數量
COL_PNO    = 0   # A: 料號
COL_NAME   = 1   # B: 品名
COL_WO     = 5   # F: 工單單號
COL_DEMAND = 7   # H: 需求數量
COL_WODATE = 13  # N: 工單開工日

# =========================
# Sidebar 設定
# =========================
with st.sidebar:
    st.divider()
    st.markdown("### ⚙️ 設定")

    shortage_file  = st.file_uploader("📂 上傳廠內排程表（工單缺料）", type=["xlsx", "xls", "csv"])
    sd_file        = st.file_uploader("📂 上傳供需表",                  type=["xlsx", "xls", "csv"])
    shipdate_file  = st.file_uploader("📂 上傳出貨日（選填）",          type=["xlsx", "xls"])

    st.markdown("**📅 分析區間**")
    date_start = st.date_input("起始日", datetime(2026, 5, 1),  format="YYYY/MM/DD")
    date_end   = st.date_input("結束日", datetime(2026, 5, 31), format="YYYY/MM/DD")

    if date_end < date_start:
        st.error("⚠️ 結束日不可早於起始日！")

    days = (date_end - date_start).days + 1
    st.caption(f"共 **{days}** 天（{date_start} ～ {date_end}）")

    st.divider()
    st.info(
        "💡 **配料邏輯**\n\n"
        "1️⃣ 從廠內排程表取得工單需求量（H欄）\n\n"
        "2️⃣ 從供需表計算六個倉庫存：\n"
        "　　電子倉／機構倉／半成品倉／成品倉\n"
        "　　／生產加工倉／包材倉\n\n"
        "3️⃣ 四倉庫存 ≥ 需求量 → **✅ 齊料**\n\n"
        "4️⃣ 四倉庫存 < 需求量 → 缺料，顯示\n"
        "　　供需表中的**預計進貨日＋數量**"
    )

# =========================
# 主畫面 - 空狀態
# =========================
if not shortage_file or not sd_file:
    st.info("👈 請在左側上傳「廠內排程表」及「供需表」開始分析")
    st.markdown("""
    <div style="background:#f0fdf4;border:1.5px dashed #86efac;border-radius:12px;padding:20px 24px;margin-top:16px;">
    <b style="color:#15803d;font-size:1rem;">📋 操作步驟</b>
    <ol style="color:#374151;margin-top:10px;line-height:2.2;">
      <li>ERP → 製令系統 → 匯出 <b>廠內排程表（工單缺料明細）</b>，上傳至左側</li>
      <li>ERP → 供需管理 → 匯出 <b>供需表（分倉）</b>，上傳至左側</li>
      <li>設定<b>分析區間</b>（起始日 ～ 結束日）</li>
    </ol>
    <br>
    <b style="color:#15803d;">🎯 判斷邏輯</b>
    <table style="margin-top:8px;width:100%;border-collapse:collapse;font-size:0.88rem;">
      <tr style="background:#dcfce7;">
        <td style="padding:6px 10px;">✅ 齊料</td>
        <td style="padding:6px 10px;">電子倉＋機構倉＋半成品倉＋成品倉＋生產加工倉＋包材倉 的庫存 ≥ 工單需求量</td>
      </tr>
      <tr>
        <td style="padding:6px 10px;">🔴 缺料</td>
        <td style="padding:6px 10px;">四倉庫存 &lt; 工單需求量，顯示缺料量及供需表中的預計進貨日期與數量</td>
      </tr>
    </table>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

with st.spinner("分析中，請稍候..."):

    # ── 讀廠內排程表（位置讀欄，避免中文亂碼）────────────────────────────────
    try:
        if shortage_file.name.lower().endswith('.csv'):
            sf_raw = None
            for enc in ['utf-8-sig', 'cp950', 'big5']:
                try:
                    sf_raw = pd.read_csv(shortage_file, header=None, encoding=enc)
                    break
                except Exception:
                    shortage_file.seek(0)
            if sf_raw is None:
                st.error("廠內排程表 CSV 無法讀取。")
                st.stop()
        else:
            try:
                sf_raw = pd.read_excel(shortage_file, sheet_name=0, header=None, engine='calamine')
            except Exception:
                shortage_file.seek(0)
                sf_raw = pd.read_excel(shortage_file, sheet_name=0, header=None, engine='openpyxl')
    except Exception as e:
        st.error(f"廠內排程表讀取失敗：{e}")
        st.stop()

    # 跳過第1列（欄名列）
    sf = sf_raw.iloc[1:].reset_index(drop=True).copy()

    if sf.shape[1] <= COL_DEMAND:
        st.error(f"廠內排程表欄位數不足（需至少 {COL_DEMAND+1} 欄，偵測到 {sf.shape[1]} 欄）")
        st.stop()

    sf['_pno']    = sf.iloc[:, COL_PNO].astype(str).str.strip()
    sf['_name']   = sf.iloc[:, COL_NAME].astype(str).str.strip() if sf.shape[1] > COL_NAME   else ''
    sf['_wo']     = sf.iloc[:, COL_WO].astype(str).str.strip()   if sf.shape[1] > COL_WO     else ''
    sf['_demand'] = pd.to_numeric(sf.iloc[:, COL_DEMAND], errors='coerce').fillna(0)
    sf['_wodate'] = sf.iloc[:, COL_WODATE].astype(str).str.strip() if sf.shape[1] > COL_WODATE else ''

    sf = sf[sf['_pno'].notna() & (sf['_pno'] != '') & (sf['_pno'] != 'nan')].copy()

    # ── 讀供需表 ──────────────────────────────────────────────────────────────
    try:
        if sd_file.name.lower().endswith('.csv'):
            sd = None
            for enc in ['utf-8-sig', 'cp950', 'big5']:
                try:
                    sd = pd.read_csv(sd_file, header=0, encoding=enc)
                    break
                except Exception:
                    sd_file.seek(0)
            if sd is None:
                st.error("供需表 CSV 無法讀取。")
                st.stop()
        else:
            try:
                sd = pd.read_excel(sd_file, sheet_name=0, header=0, engine='calamine')
            except Exception:
                sd_file.seek(0)
                sd = pd.read_excel(sd_file, sheet_name=0, header=0, engine='openpyxl')
        sd['日期'] = pd.to_datetime(sd['日期'], errors='coerce')
    except Exception as e:
        st.error(f"供需表讀取失敗：{e}")
        st.stop()

    for col in ['品號', '庫別名稱', '日期', '異動別', '異動數量']:
        if col not in sd.columns:
            st.error(f"供需表找不到「{col}」欄位，請確認檔案格式。")
            st.stop()

    end = pd.Timestamp(date_end)

    # ── 讀出貨日（選填）：C欄=工單號, V欄=備註文字 ──────────────────────────
    # 出貨日.xlsx 前3列為說明列，第3列（row index 2）為欄名，第4列起為資料
    shipdate_map = {}   # { 工單號: V欄文字 }
    if shipdate_file is not None:
        try:
            # 出貨日.xlsx 的工單資料在第二張工作表（index 1）
            sh_raw = pd.read_excel(shipdate_file, sheet_name=1, header=None, engine='openpyxl')
            # 第3列（index 2）為欄名列，從第4列（index 3）起為資料
            sh_data = sh_raw.iloc[3:].reset_index(drop=True)
            n_cols  = sh_data.shape[1]
            # C欄 = index 2, V欄 = index 21
            if n_cols > 2:
                wo_col_v = sh_data.iloc[:, 2].astype(str).str.strip()
            else:
                wo_col_v = pd.Series([''] * len(sh_data))
            if n_cols > 21:
                txt_col_v = sh_data.iloc[:, 21].astype(str).str.strip()
            else:
                txt_col_v = pd.Series([''] * len(sh_data))
                st.warning(f"出貨日欄位數不足（偵測到 {n_cols} 欄，V欄需第22欄），出貨備註將為空白。")
            raw_txt = sh_data.iloc[:, 21]  # 原始值（含 datetime）
            for wo, raw in zip(wo_col_v, raw_txt):
                if wo in ('', 'nan', 'None'): continue
                if pd.isna(raw): continue
                # datetime → 格式化為 YYYY/MM/DD
                if hasattr(raw, 'strftime'):
                    txt = raw.strftime('%Y/%m/%d')
                else:
                    txt = str(raw).strip()
                if txt not in ('', 'nan', 'None', 'NaT'):
                    shipdate_map[wo] = txt
        except Exception as e:
            st.warning(f"出貨日讀取失敗（略過）：{e}")

    # ── Helper 函數 ──────────────────────────────────────────────────────────

    def get_avail(pno, wh_code):
        """計算指定倉可用量（扣除預計進貨及預計生產）"""
        w = sd[(sd['品號'] == pno) & (sd['庫別'] == wh_code)]
        if w.empty: return 0
        dated = w[w['日期'].notna() & w['預計結存'].notna()]
        in_range = dated[dated['日期'] <= end]
        if not in_range.empty:
            last_bal   = in_range.sort_values('日期').iloc[-1]['預計結存']
            planned_in = dated[
                (dated['日期'] <= end) &
                (dated['異動別'].isin(['預計進貨', '預計生產']))
            ]['異動數量'].sum()
            return max(0, last_bal - planned_in)
        init_rows = w[w['日期'].isna()]
        if not init_rows.empty:
            qty = init_rows['異動數量'].dropna()
            if not qty.empty: return max(0, float(qty.iloc[0]))
        return 0

    def avail_4wh(pno):
        """計算 VALID_SRC 六倉的可用量合計（一般工單用）"""
        part_sd    = sd[sd['品號'] == pno]
        dated_part = part_sd[part_sd['日期'].notna() & part_sd['庫別名稱'].notna()]
        code_name  = (dated_part[['庫別', '庫別名稱']]
                      .drop_duplicates('庫別')
                      .set_index('庫別')['庫別名稱']
                      .to_dict())
        code_name  = {c: n for c, n in code_name.items() if len(str(c)) <= 12}
        names_with_dated = set(code_name.values())
        dated_codes      = set(code_name.keys())
        total = 0.0
        for wh_code, wh_name in code_name.items():
            if wh_name not in VALID_SRC: continue
            total += get_avail(pno, wh_code)
        init_rows = part_sd[part_sd['日期'].isna() & part_sd['庫別名稱'].isna()]
        for wh_k in init_rows['庫別'].dropna().unique():
            ws = str(wh_k)
            if ws in dated_codes or ws in names_with_dated: continue
            if ws not in VALID_SRC or len(ws) > 12: continue
            qty = init_rows[init_rows['庫別'] == wh_k]['異動數量'].dropna()
            if not qty.empty and float(qty.iloc[0]) > 0:
                total += float(qty.iloc[0])
        return int(total)

    def avail_prod_wh(pno):
        """5142 工單專用：直接查「加工倉」庫存（同時相容生産/生產兩種字形）"""
        part_sd = sd[sd['品號'] == pno]
        # 用「加工倉」關鍵字比對，避免簡繁體字形不同造成錯誤
        prod = part_sd[part_sd['庫別名稱'].astype(str).str.contains('加工倉', na=False)]
        if prod.empty:
            return 0
        dated = prod[prod['日期'].notna() & prod['預計結存'].notna()]
        in_range = dated[dated['日期'] <= end]
        if not in_range.empty:
            last_bal = in_range.sort_values('日期').iloc[-1]['預計結存']
            return max(0, int(last_bal))
        # 沒有日期列 → 找初始異動數量
        init_r = prod[prod['日期'].isna()]
        if not init_r.empty:
            qty = init_r['異動數量'].dropna()
            if not qty.empty:
                return max(0, int(float(qty.iloc[0])))
        return 0

    def get_other_wh_stocks(pno):
        """5142 工單：取得除加工倉外其餘 VALID_SRC 倉的庫存字串，格式：'電子倉(100)、機構倉(50)'"""
        part_sd    = sd[sd['品號'] == pno]
        # 排除生産/生產加工倉（用關鍵字，相容簡繁體）
        dated_part = part_sd[
            part_sd['日期'].notna() &
            part_sd['庫別名稱'].notna() &
            ~part_sd['庫別名稱'].astype(str).str.contains('加工倉', na=False)
        ]
        code_name  = (dated_part[['庫別', '庫別名稱']]
                      .drop_duplicates('庫別')
                      .set_index('庫別')['庫別名稱']
                      .to_dict())
        parts = []
        for wh_code, wh_name in code_name.items():
            if wh_name not in VALID_SRC: continue
            qty = int(get_avail(pno, wh_code))
            if qty > 0:
                parts.append(f'{wh_name}({qty:,})')
        return '、'.join(parts) if parts else ''

    def get_incoming(pno):
        """從供需表抓預計進貨＋預計生產：不限分析區間，依日期排序，格式 [類型]MM/DD(數量)"""
        sub = sd[
            (sd['品號'] == pno) &
            (sd['異動別'].isin(['預計進貨', '預計生產'])) &
            (sd['日期'].notna())
        ].sort_values('日期')
        if sub.empty: return None
        parts = []
        for _, row in sub.iterrows():
            tag = '進貨' if row['異動別'] == '預計進貨' else '生產'
            parts.append(f"[{tag}]{row['日期'].strftime('%m/%d')}({int(row['異動數量'])})")
        return '、'.join(parts)

    # ── 主分析 ────────────────────────────────────────────────────────────────
    # 逐列處理：每張工單每個料號各自一行
    # 5142 工單：只算生產加工倉做齊缺判斷，其餘五倉庫存顯示於預計進料欄
    rows             = []
    avail_cache      = {}    # 一般工單：六倉合計
    avail_cache_5142 = {}    # 5142 工單：只算生產加工倉
    other_wh_cache   = {}    # 5142 工單：其餘五倉庫存字串（顯示於預計進料欄）

    for _, row_sf in sf.iterrows():
        pno_str  = str(row_sf['_pno']).strip()
        demand   = int(row_sf['_demand'])
        wo       = str(row_sf['_wo'])
        wo       = '' if wo in ('nan', 'None') else wo
        wo_date  = str(row_sf['_wodate'])
        wo_date  = '' if wo_date in ('nan', 'None', 'NaT') else wo_date

        part_name = str(row_sf['_name'])
        if part_name in ('', 'nan', 'None'): part_name = ''

        if demand <= 0:
            continue

        is_5142 = wo.startswith('5142')

        if is_5142:
            # 5142：只算生産加工倉（直接查詢，避免字形問題）
            if pno_str not in avail_cache_5142:
                avail_cache_5142[pno_str] = avail_prod_wh(pno_str)
            avail = avail_cache_5142[pno_str]
            # 其餘五倉庫存字串（顯示於預計進料欄）
            if pno_str not in other_wh_cache:
                other_wh_cache[pno_str] = get_other_wh_stocks(pno_str)
        else:
            if pno_str not in avail_cache:
                avail_cache[pno_str] = avail_4wh(pno_str)
            avail = avail_cache[pno_str]

        # 5142 工單：其餘五倉庫存字串（顯示於預計進料欄）
        other_wh_str = other_wh_cache.get(pno_str, '') if is_5142 else ''

        # 判斷齊料 / 缺料
        ship_note = shipdate_map.get(wo, '')

        if avail >= demand:
            rows.append({
                '工單單號':             wo,
                '開工日':               wo_date,
                '料號':                 pno_str,
                '品名':                 part_name,
                '工單需求量':           demand,
                '六倉可用庫存':         avail,
                '缺料量':               0,
                '狀態':                 '✅ 齊料',
                '預計進料日（含數量）': other_wh_str,  # 5142：其餘五倉庫存；其餘工單為空
                '出貨備註':             ship_note,
                '_is_short':            False,
            })
        else:
            shortage = demand - avail
            if pno_str not in [r['料號'] for r in rows if r['_is_short']]:
                incoming = get_incoming(pno_str)
            else:
                incoming = next(
                    (r['預計進料日（含數量）'] for r in reversed(rows)
                     if r['料號'] == pno_str and r['_is_short']),
                    get_incoming(pno_str)
                )
            # 5142：其餘五倉庫存字串加在最前面，再接預計進貨/生產
            incoming_full = incoming or ''
            if other_wh_str:
                incoming_full = (other_wh_str + '、' + incoming_full).rstrip('、') if incoming_full else other_wh_str
            rows.append({
                '工單單號':             wo,
                '開工日':               wo_date,
                '料號':                 pno_str,
                '品名':                 part_name,
                '工單需求量':           demand,
                '六倉可用庫存':         avail,
                '缺料量':               shortage,
                '狀態':                 f'🔴 缺料 {shortage:,}',
                '預計進料日（含數量）': incoming_full or '—（供需表無預計進貨）',
                '出貨備註':             ship_note,
                '_is_short':            True,
            })

    df_out = pd.DataFrame(rows) if rows else pd.DataFrame()

# ── 統計卡片 ──────────────────────────────────────────────────────────────────
total_pno   = sf['_pno'].nunique()
n_ok        = len(df_out[df_out['_is_short'] == False]) if len(df_out) else 0
n_short     = len(df_out[df_out['_is_short'] == True])  if len(df_out) else 0

col1, col2, col3, col4 = st.columns(4)
col1.metric("工單料號總數",  f"{total_pno} 個")
col2.metric("✅ 齊料",       f"{n_ok} 個")
col3.metric("🔴 缺料",       f"{n_short} 個",
            delta=None if n_short == 0 else "需確認進貨日",
            delta_color="inverse")
col4.metric("缺料量合計",
            f"{int(df_out[df_out['_is_short']==True]['缺料量'].sum()):,}" if n_short > 0 else "0")

st.divider()
st.markdown(f"#### 🏭 廠內配料表　{date_start} ～ {date_end}")

if df_out.empty:
    st.success("✅ 沒有工單需求資料，請確認廠內排程表內容。")
else:

    display_cols = ['工單單號', '開工日', '料號', '品名', '工單需求量',
                    '六倉可用庫存', '缺料量', '狀態', '預計進料日（含數量）', '出貨備註']
    # 明細表只顯示缺料項目
    df_short   = df_out[df_out['_is_short'] == True].reset_index(drop=True)
    df_display = df_short[[c for c in display_cols if c in df_short.columns]].copy()

    # 5142 工單完整檢視（含齊料），方便確認計算是否正確
    df_5142_all = df_out[df_out['工單單號'].str.startswith('5142', na=False)].reset_index(drop=True)
    if not df_5142_all.empty:
        with st.expander(f"📋 5142 工單完整明細（含齊料，共 {len(df_5142_all)} 筆）", expanded=False):
            st.caption("5142 工單只以「生產加工倉」庫存做齊缺判斷，六倉可用庫存欄 = 生產加工倉庫存")
            def _style_5142(row):
                if '缺料' in str(row.get('狀態', '')):
                    return ['background-color:#fef2f2; color:#991b1b; font-weight:600;'] * len(row)
                return ['background-color:#f0fdf4; color:#15803d;'] * len(row)
            df_5142_disp = df_5142_all[[c for c in display_cols if c in df_5142_all.columns]].copy()
            st.dataframe(
                df_5142_disp.style.apply(_style_5142, axis=1),
                use_container_width=True,
                hide_index=True,
                column_config={
                    '預計進料日（含數量）': st.column_config.TextColumn(width='large'),
                    '品名':               st.column_config.TextColumn(width='medium'),
                    '工單單號':           st.column_config.TextColumn(width='medium'),
                }
            )

    def _row_style2(row):
        return ['background-color:#fef2f2; color:#991b1b; font-weight:600;'] * len(row)

    st.dataframe(
        df_display.style.apply(_row_style2, axis=1),
        use_container_width=True,
        height=560,
        hide_index=True,
        column_config={
            '預計進料日（含數量）': st.column_config.TextColumn(width='large'),
            '出貨備註':           st.column_config.TextColumn(width='large'),
            '品名':               st.column_config.TextColumn(width='medium'),
            '工單單號':           st.column_config.TextColumn(width='medium'),
            '狀態':               st.column_config.TextColumn(width='small'),
        }
    )

    # ── 匯出 Excel ────────────────────────────────────────────────────────────
    def build_excel(df, df_meta, start_dt, end_dt):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '廠內配料表'
        thin   = Side(style='thin', color='FFCCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers    = ['工單單號', '開工日', '料號', '品名', '工單需求量',
                      '六倉可用庫存', '缺料量', '狀態', '預計進料日（含數量）', '出貨備註']
        col_widths = [28, 14, 32, 28, 12, 14, 12, 14, 40, 36]
        hdr_colors = ['FFF2F2F2', 'FFFFF0CC', 'FFD9E8FF', 'FFF5F5F5', 'FFE8F4FD',
                      'FFE8F4FD', 'FFFCE4D6', 'FFF2F2F2', 'FFE8F4FD', 'FFF5E6FF']
        left_cols  = {1, 2, 3, 4, 9, 10}

        total_cols = len(headers)
        merge_end  = chr(64 + total_cols)
        ws.merge_cells(f'A1:{merge_end}1')
        c = ws['A1']
        c.value     = f'廠內配料表　{start_dt.strftime("%Y/%m/%d")} ～ {end_dt.strftime("%Y/%m/%d")}'
        c.font      = Font(name='Arial', bold=True, size=12, color='FFFFFFFF')
        c.fill      = PatternFill('solid', start_color='FF374151')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24

        for i, (h, hc) in enumerate(zip(headers, hdr_colors), 1):
            cell = ws.cell(row=2, column=i, value=h)
            cell.font      = Font(name='Arial', bold=True, size=9)
            cell.fill      = PatternFill('solid', start_color=hc)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = border
        ws.row_dimensions[2].height = 28

        for r_i, (row_dict, is_short) in enumerate(
                zip(df.to_dict('records'), df_meta['_is_short'].values), 3):
            for c_i, key in enumerate(headers, 1):
                val  = row_dict.get(key)
                cell = ws.cell(row=r_i, column=c_i, value=val)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='left' if c_i in left_cols else 'center',
                    vertical='center',
                    wrap_text=(c_i == 8),
                )
                if is_short:
                    cell.fill = PatternFill('solid', start_color='FFFCE4EC')
                    cell.font = Font(name='Arial', size=9, bold=True, color='FF991B1B')
                    if c_i == 8:
                        ws.row_dimensions[r_i].height = 42
                else:
                    cell.fill = PatternFill('solid', start_color='FFF0FDF4')
                    cell.font = Font(name='Arial', size=9, color='FF15803D')

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A3'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    buf = build_excel(df_display, df_short, date_start, date_end)
    st.download_button(
        label="⬇️ 匯出廠內配料表（Excel）",
        data=buf,
        file_name=f"廠內配料表_{date_start.strftime('%Y%m%d')}_{date_end.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ── 診斷查詢（開發用）─────────────────────────────────────────────────────────
with st.expander("🔍 倉庫庫存診斷查詢", expanded=False):
    st.caption("輸入工單號 + 料號，查看各倉庫存的詳細計算過程")
    diag_col1, diag_col2 = st.columns(2)
    diag_wo  = diag_col1.text_input("工單號碼", placeholder="e.g. 5142-20260317002")
    diag_pno = diag_col2.text_input("料號",     placeholder="e.g. 1903-10-01-245176-00")

    if diag_wo.strip() and diag_pno.strip():
        dpno = diag_pno.strip()
        dwo  = diag_wo.strip()
        dis_5142 = dwo.startswith('5142')

        st.markdown(f"**工單**：`{dwo}` {'→ 5142工單（只算生産加工倉）' if dis_5142 else '→ 一般工單（六倉合計）'}")

        # ① 先看廠內排程表裡有沒有這筆
        sf_match = sf[sf['_wo'] == dwo]
        pno_match = sf_match[sf_match['_pno'] == dpno]
        if pno_match.empty:
            st.warning(f"⚠️ 廠內排程表中找不到 工單={dwo} 且 料號={dpno} 的資料")
            # 列出該工單下所有料號
            if not sf_match.empty:
                st.write("該工單下的所有料號：", sf_match['_pno'].tolist())
        else:
            demand_diag = int(pno_match.iloc[0]['_demand'])
            st.info(f"📋 廠內排程需求量：**{demand_diag:,}**")

        # ② 供需表裡查該料號的所有倉別
        part_sd_diag = sd[sd['品號'] == dpno]
        if part_sd_diag.empty:
            st.error(f"供需表中找不到料號 {dpno}")
        else:
            dated_diag = part_sd_diag[part_sd_diag['日期'].notna() & part_sd_diag['庫別名稱'].notna()]
            code_name_diag = (dated_diag[['庫別', '庫別名稱']]
                              .drop_duplicates('庫別')
                              .set_index('庫別')['庫別名稱']
                              .to_dict())
            code_name_diag = {c: n for c, n in code_name_diag.items() if len(str(c)) <= 12}

            st.markdown("**各倉庫存計算明細：**")
            diag_rows = []
            total_all = 0
            total_5142 = 0   # 只含生産加工倉
            for wh_code, wh_name in code_name_diag.items():
                in_valid = wh_name in VALID_SRC
                qty = int(get_avail(dpno, wh_code)) if in_valid else 0

                # 最後一筆結存 & planned_in（透明化）
                w_rows = part_sd_diag[part_sd_diag['庫別'] == wh_code]
                dated_w = w_rows[w_rows['日期'].notna() & w_rows['預計結存'].notna()]
                in_range_w = dated_w[dated_w['日期'] <= end]
                if not in_range_w.empty:
                    last_bal_w = in_range_w.sort_values('日期').iloc[-1]['預計結存']
                    planned_w  = dated_w[
                        (dated_w['日期'] <= end) &
                        (dated_w['異動別'].isin(['預計進貨', '預計生産']))
                    ]['異動數量'].sum()
                    detail_str = f"預計結存={int(last_bal_w):,}  -  預計進貨/生産={int(planned_w):,}  =  可用{qty:,}"
                else:
                    detail_str = "無日期資料"

                use_for_5142 = (wh_name == '生産加工倉')
                diag_rows.append({
                    '倉別代碼': str(wh_code),
                    '倉別名稱': wh_name,
                    '在VALID_SRC': '✅' if in_valid else '❌',
                    '5142計算': '✅ 計入' if use_for_5142 else '—',
                    '可用庫存': qty,
                    '計算明細': detail_str,
                })
                if in_valid:
                    total_all += qty
                if use_for_5142:
                    total_5142 += qty

            # init_rows（日期＆倉別名稱均為空）
            init_diag = part_sd_diag[part_sd_diag['日期'].isna() & part_sd_diag['庫別名稱'].isna()]
            dated_codes_diag = set(code_name_diag.keys())
            names_diag = set(code_name_diag.values())
            for wh_k in init_diag['庫別'].dropna().unique():
                ws2 = str(wh_k)
                if ws2 in dated_codes_diag or ws2 in names_diag:
                    continue
                in_valid2 = ws2 in VALID_SRC and len(ws2) <= 12
                qty2_raw = init_diag[init_diag['庫別'] == wh_k]['異動數量'].dropna()
                qty2 = int(float(qty2_raw.iloc[0])) if not qty2_raw.empty and float(qty2_raw.iloc[0]) > 0 else 0
                use_for_5142_2 = (ws2 == '生産加工倉')
                diag_rows.append({
                    '倉別代碼': ws2,
                    '倉別名稱': '(init_rows，無名稱)',
                    '在VALID_SRC': '✅' if in_valid2 else '❌',
                    '5142計算': '✅ 計入' if use_for_5142_2 else '—',
                    '可用庫存': qty2 if in_valid2 else 0,
                    '計算明細': f'init異動數量={qty2}',
                })
                if in_valid2:
                    total_all += qty2
                if use_for_5142_2:
                    total_5142 += qty2

            df_diag = pd.DataFrame(diag_rows)
            st.dataframe(df_diag, use_container_width=True, hide_index=True)
            st.markdown(
                f"**六倉合計可用庫存（一般工單）：{total_all:,}**　｜　"
                f"**生産加工倉可用庫存（5142工單）：{total_5142:,}**"
            )
            if not pno_match.empty:
                demand_diag = int(pno_match.iloc[0]['_demand'])
                used = total_5142 if dis_5142 else total_all
                st.markdown(
                    f"需求量={demand_diag:,}　可用={used:,}　"
                    + ("→ **應為 ✅ 齊料**" if used >= demand_diag else "→ **應為 🔴 缺料**")
                )
