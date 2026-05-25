import streamlit as st
import pandas as pd
import io
import sys
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from utils.shared import ensure_calamine, inject_css, render_header, render_sidebar, render_sd_loader, render_nas_loader

ensure_calamine()

if "lang" not in st.session_state:
    st.session_state["lang"] = "zh"

st.set_page_config(page_title="工單進度表", page_icon="📋", layout="wide", initial_sidebar_state="expanded")
inject_css()

with st.sidebar:
    col_zh, col_en = st.columns(2)
    with col_zh:
        if st.button("🇹🇼 中文", key="btn_zh",
                     type="primary" if st.session_state["lang"] == "zh" else "secondary",
                     use_container_width=True):
            st.session_state["lang"] = "zh"
            st.rerun()
    with col_en:
        if st.button("🇺🇸 EN", key="btn_en",
                     type="primary" if st.session_state["lang"] == "en" else "secondary",
                     use_container_width=True):
            st.session_state["lang"] = "en"
            st.rerun()

render_header(
    title="工單進度表",
    subtitle="Work Order Material Shortage Analysis &nbsp;·&nbsp; ORing Industrial Networking",
    badge="Production Control · PC",
    show_logo=False,
)
render_sidebar()

# ── 操作 SOP ─────────────────────────────────────────────────────────────────
with st.expander("📖 操作 SOP　－　點此展開", expanded=False):
    st.markdown("""
<div style="line-height:2; font-size:0.93rem; color:#1e293b;">

<b style="font-size:1.05rem; color:#1d4ed8;">📋 工單進度表　操作說明</b>

<hr style="margin:8px 0; border-color:#e2e8f0;">

**Step 1 ── 準備檔案**

| 檔案 | 來源 | 說明 |
|---|---|---|
| 廠內排程表 | 生管提供 | 含工單缺料明細（欠料數量）|
| 供需表 | 自動載入 ✅ | NAS 每日自動抓取，無需手動上傳 |
| 出貨日 | 自動載入 ✅ | 自動抓取最新寶橋早會資料 |
| 生產開完工表 | 手動上傳（選填）| 帶入預計產量、已生產量、完工日 |

---

**Step 2 ── 上傳廠內排程表**

> 左側側邊欄 → 點「📂 上傳廠內排程表」→ 選取檔案

---

**Step 3 ── 設定分析區間**

> 左側側邊欄 → 設定「起始日」與「結束日」
> ⚠️ 區間會影響工單統計筆數與庫存可用量計算

---

**Step 4 ── 查看結果**

- 上方統計卡片：區間工單總筆數（廠內 / 委外）、齊料工單、缺料工單
- 下方 Tab 分頁：依單別分類顯示缺料明細
  - 📋 **全部缺料** → 所有缺料工單一覽
  - 各 **單別 Tab** → 如 5141 廠內量產製令、5145 託外量產製令

---

**Step 5 ── 匯出報表**

> 頁面下方 → 點「⬇️ 匯出工單進度表（Excel）」

---

**💡 缺料判斷邏輯**

| 工單類型 | 庫存計算方式 |
|---|---|
| 一般工單（非5142）| 電子倉＋機構倉＋半成品倉＋成品倉＋生產加工倉＋包材倉 合計 |
| 5142 改機工單 | 僅計算生產加工倉，其餘倉顯示於「預計進料日」欄 |

- 庫存 ≥ 需求 → **✅ 齊料**
- 庫存 < 需求 → **🔴 缺料**（顯示預計進貨日期）

</div>
""", unsafe_allow_html=True)

# 六倉範圍
VALID_SRC = {'電子倉', '機構倉', '半成品倉', '成品倉', '生產加工倉', '包材倉'}

# 單別對應表（工單號前4碼 → 單據名稱）
SINGLE_TYPE_MAP = {
    '5140': '廠內試產製令',
    '5141': '廠內量產製令',
    '5142': '廠內改機製令',
    '5143': '託外打樣製令',
    '5144': '託外試產製令',
    '5145': '託外量產製令',
    '5146': '研發樣機製令',
    '5220': '託外重工製令',
    '5230': 'ECN重工製令',
    'FF01': '備料製令',
    'FF02': '打樣備料製令',
    'MO01': '熱銷備庫製令（託外）',
    'MO02': '熱銷備庫製令（廠內）',
}

def get_single_type(wo: str) -> tuple[str, str]:
    """回傳 (單別代號, 單別名稱)"""
    if not wo or wo in ('nan', 'None', ''):
        return '其他', '其他'
    prefix = wo[:4].upper()
    name = SINGLE_TYPE_MAP.get(prefix)
    if name:
        return prefix, name
    # 取SN碼專用：無固定前綴，以關鍵字判斷
    if 'SN' in wo.upper():
        return 'SN', '取SN碼專用'
    return prefix, f'其他（{prefix}）'

# 廠內排程.xlsx 欄位位置（0-indexed，跳過第1列欄名列）
COL_PNO    = 0   # A: 料號
COL_NAME   = 1   # B: 品名
COL_WO     = 5   # F: 工單單號
COL_DEMAND = 7   # H: 欠料數量
COL_WODATE = 13  # N: 工單開工日

# ── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.divider()
    st.markdown("### ⚙️ 設定")

    shortage_file   = st.file_uploader("📂 上傳廠內排程表（工單缺料）", type=["xlsx", "xls", "csv"])
    sd_source       = render_sd_loader(key="wo_progress")
    production_file = st.file_uploader("📂 上傳生產開完工表（選填）", type=["xlsx", "xls"])

    _NAS_SHIPDATE_DIR = "//192.168.2.34/MO_Storage/ORing MO/ORing-MO 工作/生管部/09. 廠內改機排程/2025早會"
    _NAS_SHIPDATE_PFX = "寶橋早會資料"
    shipdate_file = render_nas_loader(
        key="wo_progress_shipdate",
        nas_dir=_NAS_SHIPDATE_DIR,
        prefix=_NAS_SHIPDATE_PFX,
        label="📂 上傳出貨日（選填覆蓋）",
        types=["xlsx", "xls"],
    )

    st.markdown("**📅 分析區間**")
    date_start = st.date_input("起始日", datetime(2026, 5, 1),  format="YYYY/MM/DD")
    date_end   = st.date_input("結束日", datetime(2026, 5, 31), format="YYYY/MM/DD")

    if date_end < date_start:
        st.error("⚠️ 結束日不可早於起始日！")

    days = (date_end - date_start).days + 1
    st.caption(f"共 **{days}** 天（{date_start} ～ {date_end}）")

    st.divider()
    st.info(
        "💡 **配料邏輯**\n\n"
        "1️⃣ 廠內排程 H欄 = 欠料需求量\n\n"
        "2️⃣ 一般工單：電子倉＋機構倉＋半成品倉\n"
        "　　＋成品倉＋生產加工倉＋包材倉\n\n"
        "3️⃣ 5142工單：只算生產加工倉庫存\n"
        "　　其餘倉庫存量顯示於預計進料日欄\n\n"
        "4️⃣ 庫存 ≥ 需求 → **✅ 齊料**\n\n"
        "5️⃣ 庫存 < 需求 → 缺料，顯示進貨日期"
    )

# ── 空狀態 ────────────────────────────────────────────────────────────────────
if not shortage_file or not sd_source:
    st.info("👈 請在左側上傳「廠內排程表」，並載入或上傳「供需表」開始分析")
    st.stop()

with st.spinner("分析中，請稍候..."):

    # ── 讀廠內排程表 ──────────────────────────────────────────────────────────
    try:
        if shortage_file.name.lower().endswith('.csv'):
            sf_raw = None
            for enc in ['utf-8-sig', 'cp950', 'big5']:
                try:
                    sf_raw = pd.read_csv(shortage_file, header=None, encoding=enc)
                    break
                except Exception:
                    shortage_file.seek(0)
            if sf_raw is None:
                st.error("廠內排程表 CSV 無法讀取。"); st.stop()
        else:
            try:
                sf_raw = pd.read_excel(shortage_file, sheet_name=0, header=None, engine='calamine')
            except Exception:
                shortage_file.seek(0)
                sf_raw = pd.read_excel(shortage_file, sheet_name=0, header=None, engine='openpyxl')
    except Exception as e:
        st.error(f"廠內排程表讀取失敗：{e}"); st.stop()

    sf = sf_raw.iloc[1:].reset_index(drop=True).copy()

    if sf.shape[1] <= COL_DEMAND:
        st.error(f"廠內排程表欄位數不足（需至少 {COL_DEMAND+1} 欄，偵測到 {sf.shape[1]} 欄）"); st.stop()

    def _fmt_date_cell(val):
        """將 Excel 日期儲存格統一轉成 YYYY/MM/DD 字串"""
        s = str(val).strip()
        if s in ('', 'nan', 'None', 'NaT'): return ''
        try:
            return pd.to_datetime(s).strftime('%Y/%m/%d')
        except Exception:
            return s

    sf['_pno']    = sf.iloc[:, COL_PNO].astype(str).str.strip()
    sf['_name']   = sf.iloc[:, COL_NAME].astype(str).str.strip() if sf.shape[1] > COL_NAME   else ''
    sf['_wo']     = sf.iloc[:, COL_WO].astype(str).str.strip()   if sf.shape[1] > COL_WO     else ''
    sf['_demand'] = pd.to_numeric(sf.iloc[:, COL_DEMAND], errors='coerce').fillna(0)
    sf['_wodate'] = sf.iloc[:, COL_WODATE].apply(_fmt_date_cell) if sf.shape[1] > COL_WODATE else ''
    sf = sf[sf['_pno'].notna() & (sf['_pno'] != '') & (sf['_pno'] != 'nan')].copy()

    # ── 讀供需表（支援上傳檔案 or NAS 路徑字串）──────────────────────────────
    try:
        src_name = sd_source if isinstance(sd_source, str) else getattr(sd_source, 'name', '')
        if str(src_name).lower().endswith('.csv'):
            sd = None
            for enc in ['utf-8-sig', 'cp950', 'big5']:
                try:
                    sd = pd.read_csv(sd_source, header=0, encoding=enc)
                    break
                except Exception:
                    if hasattr(sd_source, 'seek'): sd_source.seek(0)
            if sd is None:
                st.error("供需表 CSV 無法讀取。"); st.stop()
        else:
            try:
                sd = pd.read_excel(sd_source, sheet_name=0, header=0, engine='calamine')
            except Exception:
                if hasattr(sd_source, 'seek'): sd_source.seek(0)
                sd = pd.read_excel(sd_source, sheet_name=0, header=0, engine='openpyxl')
        sd['日期'] = pd.to_datetime(sd['日期'], errors='coerce')
    except Exception as e:
        st.error(f"供需表讀取失敗：{e}"); st.stop()

    for col in ['品號', '庫別', '庫別名稱', '日期', '異動別', '異動數量', '預計結存']:
        if col not in sd.columns:
            st.error(f"供需表找不到「{col}」欄位，請確認檔案格式。"); st.stop()

    end   = pd.Timestamp(date_end)
    start = pd.Timestamp(date_start)

    # ── 預建品號索引（大幅加速：避免主迴圈每筆重複全表篩選）────────────────────
    sd_by_pno: dict = {pno: grp for pno, grp in sd.groupby('品號', sort=False)}

    # ── 讀生產開完工表（選填）：A欄=製令單號, F欄=完工日, L欄=預計產量, N欄=已生產量 ──
    prod_map = {}   # {製令單號: {'完工日':..., '預計產量':..., '已生產量':...}}
    _prod_src = production_file
    if _prod_src is None and os.path.exists(r'C:\Users\T26019\Desktop\生產開完工表.xlsx'):
        _prod_src = r'C:\Users\T26019\Desktop\生產開完工表.xlsx'
    if _prod_src is not None:
        try:
            _pf = open(_prod_src, 'rb') if isinstance(_prod_src, str) else _prod_src
            pf_df = pd.read_excel(_pf, header=0, engine='openpyxl', dtype=str)
            if isinstance(_prod_src, str): _pf.close()
            pf_cols = list(pf_df.columns)
            # 取欄位（以位置取，不依賴欄名）
            _wo_col   = pf_df.iloc[:, 0].astype(str).str.strip()   # A: 製令單號
            _done_col = pf_df.iloc[:, 5].astype(str).str.strip()   # F: 完工日
            _plan_col = pf_df.iloc[:, 11]                           # L: 預計產量
            _real_col = pf_df.iloc[:, 13]                           # N: 已生產量
            for wo_no, done, plan, real in zip(_wo_col, _done_col, _plan_col, _real_col):
                if not wo_no or wo_no in ('nan', 'None', ''): continue
                prod_map[wo_no] = {
                    '完工日':   '' if done in ('nan', 'None', 'NaT') else done,
                    '預計產量': int(float(plan)) if str(plan) not in ('nan', 'None', '') else 0,
                    '已生產量': int(float(real)) if str(real) not in ('nan', 'None', '') else 0,
                }
        except Exception as e:
            st.warning(f"生產開完工表讀取失敗（略過）：{e}")

    # ── 讀出貨日（選填）：sheet=1, 跳3列, C欄=工單號, V欄=備註 ──────────────
    shipdate_map = {}
    if shipdate_file is not None:
        try:
            _sd_src = shipdate_file if not isinstance(shipdate_file, str) else open(shipdate_file, 'rb')
            sh_raw  = pd.read_excel(_sd_src, sheet_name=1, header=None, engine='openpyxl')
            if isinstance(shipdate_file, str): _sd_src.close()
            sh_data = sh_raw.iloc[3:].reset_index(drop=True)
            n_cols  = sh_data.shape[1]
            wo_col  = sh_data.iloc[:, 2].astype(str).str.strip() if n_cols > 2  else pd.Series([''] * len(sh_data))
            raw_txt = sh_data.iloc[:, 21]                                         if n_cols > 21 else pd.Series([None]  * len(sh_data))
            for wo, raw in zip(wo_col, raw_txt):
                if wo in ('', 'nan', 'None'): continue
                if pd.isna(raw): continue
                txt = raw.strftime('%Y/%m/%d') if hasattr(raw, 'strftime') else str(raw).strip()
                if txt not in ('', 'nan', 'None', 'NaT'):
                    shipdate_map[wo] = txt
        except Exception as e:
            st.warning(f"出貨日讀取失敗（略過）：{e}")

    # ── Helper 函數（全部改用 sd_by_pno 索引，避免重複全表掃描）────────────────

    def get_avail(part_sd, wh_code):
        """指定倉可用量，傳入已篩好的品號子表"""
        w = part_sd[part_sd['庫別'] == wh_code]
        if w.empty: return 0
        dated    = w[w['日期'].notna() & w['預計結存'].notna()]
        in_range = dated[dated['日期'] <= end]
        if not in_range.empty:
            last_bal   = in_range.sort_values('日期').iloc[-1]['預計結存']
            planned_in = dated[
                (dated['日期'] <= end) &
                (dated['異動別'].isin(['預計進貨', '預計生產']))
            ]['異動數量'].sum()
            return max(0, last_bal - planned_in)
        init_r = w[w['日期'].isna()]
        if not init_r.empty:
            qty = init_r['異動數量'].dropna()
            if not qty.empty: return max(0, float(qty.iloc[0]))
        return 0

    def avail_4wh(pno):
        part_sd    = sd_by_pno.get(pno, pd.DataFrame(columns=sd.columns))
        if part_sd.empty: return 0
        dated_part = part_sd[part_sd['日期'].notna() & part_sd['庫別名稱'].notna()]
        code_name  = (dated_part[['庫別', '庫別名稱']]
                      .drop_duplicates('庫別')
                      .set_index('庫別')['庫別名稱']
                      .to_dict())
        code_name        = {c: n for c, n in code_name.items() if len(str(c)) <= 12}
        names_with_dated = set(code_name.values())
        dated_codes      = set(code_name.keys())
        total = 0.0
        for wh_code, wh_name in code_name.items():
            if wh_name not in VALID_SRC: continue
            total += get_avail(part_sd, wh_code)
        init_rows = part_sd[part_sd['日期'].isna() & part_sd['庫別名稱'].isna()]
        for wh_k in init_rows['庫別'].dropna().unique():
            ws = str(wh_k)
            if ws in dated_codes or ws in names_with_dated: continue
            if ws not in VALID_SRC or len(ws) > 12: continue
            qty = init_rows[init_rows['庫別'] == wh_k]['異動數量'].dropna()
            if not qty.empty and float(qty.iloc[0]) > 0:
                total += float(qty.iloc[0])
        return int(total)

    def avail_prod_wh(pno):
        part_sd = sd_by_pno.get(pno, pd.DataFrame(columns=sd.columns))
        if part_sd.empty: return 0
        prod    = part_sd[part_sd['庫別名稱'].astype(str).str.contains('加工倉', na=False)]
        if prod.empty: return 0
        dated = prod[prod['日期'].notna() & prod['預計結存'].notna()]
        before_start = dated[dated['日期'] < start]
        if not before_start.empty:
            opening = int(before_start.sort_values('日期').iloc[-1]['預計結存'])
        else:
            in_range_all = dated[dated['日期'] <= end]
            if in_range_all.empty:
                init_r = prod[prod['日期'].isna()]
                if not init_r.empty:
                    qty = init_r['異動數量'].dropna()
                    if not qty.empty: return max(0, int(float(qty.iloc[0])))
                return 0
            first   = in_range_all.sort_values('日期').iloc[0]
            opening = int(first['預計結存']) + int(first['異動數量'])
        in_period  = dated[(dated['日期'] >= start) & (dated['日期'] <= end)]
        planned_in = int(in_period[
            in_period['異動別'].isin(['預計進貨', '預計生產'])
        ]['異動數量'].sum())
        return max(0, opening + planned_in)

    def get_other_wh_stocks(pno):
        part_sd    = sd_by_pno.get(pno, pd.DataFrame(columns=sd.columns))
        if part_sd.empty: return ''
        dated_part = part_sd[
            part_sd['日期'].notna() &
            part_sd['庫別名稱'].notna() &
            ~part_sd['庫別名稱'].astype(str).str.contains('加工倉', na=False)
        ]
        code_name        = (dated_part[['庫別', '庫別名稱']]
                            .drop_duplicates('庫別')
                            .set_index('庫別')['庫別名稱']
                            .to_dict())
        code_name        = {c: n for c, n in code_name.items() if len(str(c)) <= 12}
        names_with_dated = set(code_name.values())
        dated_codes      = set(code_name.keys())
        parts = []
        for wh_code, wh_name in code_name.items():
            if wh_name not in VALID_SRC: continue
            qty = int(get_avail(part_sd, wh_code))
            if qty > 0:
                parts.append(f'{wh_name}({qty:,})')
        init_rows = part_sd[part_sd['日期'].isna() & part_sd['庫別名稱'].isna()]
        for wh_k in init_rows['庫別'].dropna().unique():
            ws = str(wh_k)
            if ws in dated_codes or ws in names_with_dated: continue
            if ws not in VALID_SRC or len(ws) > 12: continue
            if '加工倉' in ws: continue
            qty_raw = init_rows[init_rows['庫別'] == wh_k]['異動數量'].dropna()
            if not qty_raw.empty and float(qty_raw.iloc[0]) > 0:
                parts.append(f'{ws}({int(float(qty_raw.iloc[0])):,})')
        return '、'.join(parts) if parts else ''

    def get_incoming(pno):
        part_sd = sd_by_pno.get(pno, pd.DataFrame(columns=sd.columns))
        if part_sd.empty: return None
        sub = part_sd[
            part_sd['異動別'].isin(['預計進貨', '預計生產']) &
            part_sd['日期'].notna()
        ].sort_values('日期')
        if sub.empty: return None
        parts = []
        for _, row in sub.iterrows():
            tag = '進貨' if row['異動別'] == '預計進貨' else '生產'
            parts.append(f"[{tag}]{row['日期'].strftime('%m/%d')}({int(row['異動數量'])})")
        return '、'.join(parts)

    # ── 主分析 ────────────────────────────────────────────────────────────────
    rows             = []
    avail_cache      = {}
    avail_cache_5142 = {}
    other_wh_cache   = {}
    incoming_cache   = {}

    for _, row_sf in sf.iterrows():
        pno_str   = str(row_sf['_pno']).strip()
        demand    = int(row_sf['_demand'])
        wo        = str(row_sf['_wo'])
        wo        = '' if wo in ('nan', 'None') else wo
        wo_date   = str(row_sf['_wodate'])
        wo_date   = '' if wo_date in ('nan', 'None', 'NaT') else wo_date
        part_name = str(row_sf['_name'])
        if part_name in ('', 'nan', 'None'): part_name = ''

        if demand <= 0:
            continue

        is_5142 = wo.startswith('5142')

        if is_5142:
            if pno_str not in avail_cache_5142:
                avail_cache_5142[pno_str] = avail_prod_wh(pno_str)
            avail = avail_cache_5142[pno_str]
            if pno_str not in other_wh_cache:
                other_wh_cache[pno_str] = get_other_wh_stocks(pno_str)
            other_wh_str = other_wh_cache[pno_str]
        else:
            if pno_str not in avail_cache:
                avail_cache[pno_str] = avail_4wh(pno_str)
            avail = avail_cache[pno_str]
            other_wh_str = ''

        ship_note = shipdate_map.get(wo, '')

        single_code, single_name = get_single_type(wo)
        prod_info  = prod_map.get(wo, {})
        plan_qty   = prod_info.get('預計產量', '')
        real_qty   = prod_info.get('已生產量', '')
        done_date  = prod_info.get('完工日', '')

        if avail >= demand:
            rows.append({
                '單別':                 single_code,
                '單別名稱':             single_name,
                '工單單號':             wo,
                '開工日':               wo_date,
                '完工日':               done_date,
                '預計產量':             plan_qty,
                '已生產量':             real_qty,
                '料號':                 pno_str,
                '品名':                 part_name,
                '工單需求量':           demand,
                '加工倉庫存量':         avail,
                '缺料量':               0,
                '狀態':                 '✅ 齊料',
                '預計進料日（含數量）': other_wh_str,
                '出貨備註':             ship_note,
                '_is_short':            False,
            })
        else:
            shortage = demand - avail
            if pno_str not in incoming_cache:
                incoming_cache[pno_str] = get_incoming(pno_str)
            incoming      = incoming_cache[pno_str]
            incoming_full = incoming or ''
            if other_wh_str:
                incoming_full = (other_wh_str + '、' + incoming_full).rstrip('、') if incoming_full else other_wh_str
            rows.append({
                '單別':                 single_code,
                '單別名稱':             single_name,
                '工單單號':             wo,
                '開工日':               wo_date,
                '完工日':               done_date,
                '預計產量':             plan_qty,
                '已生產量':             real_qty,
                '料號':                 pno_str,
                '品名':                 part_name,
                '工單需求量':           demand,
                '加工倉庫存量':         avail,
                '缺料量':               shortage,
                '狀態':                 f'🔴 缺料 {shortage:,}',
                '預計進料日（含數量）': incoming_full or '—（供需表無預計進貨）',
                '出貨備註':             ship_note,
                '_is_short':            True,
            })

    df_out = pd.DataFrame(rows) if rows else pd.DataFrame()

# ── 統計卡片（以工單為單位）─────────────────────────────────────────────────
# 廠內單別
_INHOUSE = {'5140', '5141', '5142', '5146', 'MO02'}

if not df_out.empty:
    _ds = pd.Timestamp(date_start)
    _de = pd.Timestamp(date_end)

    def _in_range(wo_date_str):
        if not wo_date_str or wo_date_str in ('', 'nan', 'None', 'NaT'):
            return True
        try:
            dt = pd.to_datetime(wo_date_str)
            return _ds <= dt <= _de
        except Exception:
            return True

    wo_info = df_out.drop_duplicates('工單單號')[['工單單號', '單別', '開工日']].copy()
    wo_info['_in_range'] = wo_info['開工日'].apply(_in_range)
    wo_in_range = wo_info[wo_info['_in_range']]['工單單號'].tolist()

    df_in_range  = df_out[df_out['工單單號'].isin(wo_in_range)]
    wo_short_map = df_in_range.groupby('工單單號')['_is_short'].any()
    wo_code_map  = df_in_range.drop_duplicates('工單單號').set_index('工單單號')['單別']

    all_wo       = wo_short_map.index
    ok_wo        = wo_short_map[~wo_short_map].index.tolist()
    short_wo     = wo_short_map[wo_short_map].index.tolist()

    # 廠內/委外拆分
    n_total          = len(all_wo)
    n_inhouse        = sum(1 for w in all_wo      if wo_code_map.get(w, '') in _INHOUSE)
    n_outsource      = sum(1 for w in all_wo      if wo_code_map.get(w, '') not in _INHOUSE)
    n_ok_wo          = len(ok_wo)
    n_ok_inhouse     = sum(1 for w in ok_wo       if wo_code_map.get(w, '') in _INHOUSE)
    n_ok_outsource   = sum(1 for w in ok_wo       if wo_code_map.get(w, '') not in _INHOUSE)
    n_short_wo       = len(short_wo)
    n_short_inhouse  = sum(1 for w in short_wo    if wo_code_map.get(w, '') in _INHOUSE)
    n_short_outsource= sum(1 for w in short_wo    if wo_code_map.get(w, '') not in _INHOUSE)
else:
    n_total = n_inhouse = n_outsource = 0
    n_ok_wo = n_ok_inhouse = n_ok_outsource = 0
    n_short_wo = n_short_inhouse = n_short_outsource = 0

col1, col2, col3 = st.columns(3)

with col1:
    st.metric("📋 區間工單總筆數", f"{n_total} 張")
    st.caption(f"廠內 **{n_inhouse}** 張　｜　委外 **{n_outsource}** 張")

with col2:
    st.metric("✅ 齊料工單", f"{n_ok_wo} 張")
    st.caption(f"廠內 **{n_ok_inhouse}** 張　｜　委外 **{n_ok_outsource}** 張")

with col3:
    st.metric("🔴 缺料工單", f"{n_short_wo} 張")
    st.caption(f"廠內 **{n_short_inhouse}** 張　｜　委外 **{n_short_outsource}** 張")

st.divider()
st.markdown(f"#### 📋 工單進度表　{date_start} ～ {date_end}")

if df_out.empty:
    st.success("✅ 沒有工單需求資料，請確認廠內排程表內容。")
else:
    display_cols = ['單別', '單別名稱', '工單單號', '開工日', '完工日', '預計產量', '已生產量',
                    '料號', '品名', '工單需求量', '加工倉庫存量', '缺料量', '狀態',
                    '預計進料日（含數量）', '出貨備註']

    df_short = df_out[df_out['_is_short'] == True].reset_index(drop=True)

    col_cfg = {
        '單別':               st.column_config.TextColumn(width='small'),
        '單別名稱':           st.column_config.TextColumn(width='medium'),
        '工單單號':           st.column_config.TextColumn(width='medium'),
        '開工日':             st.column_config.TextColumn(width='small'),
        '完工日':             st.column_config.TextColumn(width='small'),
        '預計產量':           st.column_config.NumberColumn(width='small'),
        '已生產量':           st.column_config.NumberColumn(width='small'),
        '料號':               st.column_config.TextColumn(width='medium'),
        '品名':               st.column_config.TextColumn(width='medium'),
        '工單需求量':         st.column_config.NumberColumn(width='small'),
        '加工倉庫存量':       st.column_config.NumberColumn(width='small'),
        '缺料量':             st.column_config.NumberColumn(width='small'),
        '狀態':               st.column_config.TextColumn(width='small'),
        '預計進料日（含數量）': st.column_config.TextColumn(width='large'),
        '出貨備註':           st.column_config.TextColumn(width='large'),
    }

    def _row_style(row):
        return ['background-color:#fef2f2; color:#991b1b; font-weight:600;'] * len(row)

    # ── 按單別分 Tab 顯示 ─────────────────────────────────────────────────────
    # 取出有缺料的單別（依照 SINGLE_TYPE_MAP 順序排序）
    order = list(SINGLE_TYPE_MAP.keys()) + ['SN', '其他']
    present_codes = df_short['單別'].unique().tolist()
    sorted_codes  = sorted(present_codes, key=lambda x: order.index(x) if x in order else 999)

    # 建立 Tab 標籤：「全部」+ 各單別
    tab_labels = ['📋 全部缺料']
    for code in sorted_codes:
        name  = SINGLE_TYPE_MAP.get(code, code)
        cnt   = (df_short['單別'] == code).sum()
        tab_labels.append(f'{code} {name}（{cnt}）')

    tabs = st.tabs(tab_labels)

    # Tab 0：全部缺料
    with tabs[0]:
        df_disp_all = df_short[[c for c in display_cols if c in df_short.columns]].copy()
        st.dataframe(
            df_disp_all.style.apply(_row_style, axis=1),
            use_container_width=True, height=520, hide_index=True, column_config=col_cfg,
        )

    # Tab 1+：各單別
    for i, code in enumerate(sorted_codes, 1):
        with tabs[i]:
            df_grp = df_short[df_short['單別'] == code]
            df_grp = df_grp[[c for c in display_cols if c in df_grp.columns]].copy()
            name   = SINGLE_TYPE_MAP.get(code, code)
            st.markdown(f"**{code} {name}　缺料 {len(df_grp)} 筆**")
            st.dataframe(
                df_grp.style.apply(_row_style, axis=1),
                use_container_width=True, height=min(520, 60 + len(df_grp) * 38),
                hide_index=True, column_config=col_cfg,
            )

    # ── 匯出 Excel ────────────────────────────────────────────────────────────
    def build_excel(df, df_meta, start_dt, end_dt):
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = '工單進度表'
        thin   = Side(style='thin', color='FFCCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers    = ['單別', '單別名稱', '工單單號', '開工日', '料號', '品名', '工單需求量',
                      '加工倉庫存量', '缺料量', '狀態', '預計進料日（含數量）', '出貨備註']
        col_widths = [10, 22, 28, 14, 32, 28, 12, 14, 12, 14, 44, 36]
        hdr_colors = ['FFE2E8F0', 'FFE2E8F0', 'FFF2F2F2', 'FFFFF0CC', 'FFD9E8FF', 'FFF5F5F5',
                      'FFE8F4FD', 'FFE8F4FD', 'FFFCE4D6', 'FFF2F2F2', 'FFE8F4FD', 'FFF5E6FF']
        left_cols  = {1, 2, 3, 4, 5, 6, 11, 12}

        total_cols = len(headers)
        merge_end  = chr(64 + total_cols)
        ws.merge_cells(f'A1:{merge_end}1')
        c           = ws['A1']
        c.value     = f'工單進度表　{start_dt.strftime("%Y/%m/%d")} ～ {end_dt.strftime("%Y/%m/%d")}'
        c.font      = Font(name='Arial', bold=True, size=12, color='FFFFFFFF')
        c.fill      = PatternFill('solid', start_color='FF374151')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24

        for i, (h, hc) in enumerate(zip(headers, hdr_colors), 1):
            cell            = ws.cell(row=2, column=i, value=h)
            cell.font       = Font(name='Arial', bold=True, size=9)
            cell.fill       = PatternFill('solid', start_color=hc)
            cell.alignment  = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border     = border
        ws.row_dimensions[2].height = 28

        for r_i, (row_dict, is_short) in enumerate(
                zip(df.to_dict('records'), df_meta['_is_short'].values), 3):
            for c_i, key in enumerate(headers, 1):
                val  = row_dict.get(key)
                cell = ws.cell(row=r_i, column=c_i, value=val)
                cell.border    = border
                cell.alignment = Alignment(
                    horizontal='left' if c_i in left_cols else 'center',
                    vertical='center',
                    wrap_text=(c_i in {9, 10}),
                )
                if is_short:
                    cell.fill = PatternFill('solid', start_color='FFFCE4EC')
                    cell.font = Font(name='Arial', size=9, bold=True, color='FF991B1B')
                    if c_i in {9, 10}:
                        ws.row_dimensions[r_i].height = 42
                else:
                    cell.fill = PatternFill('solid', start_color='FFF0FDF4')
                    cell.font = Font(name='Arial', size=9, color='FF15803D')

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A3'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    df_display = df_short[[c for c in display_cols if c in df_short.columns]].copy()
    buf = build_excel(df_display, df_short, date_start, date_end)
    st.download_button(
        label="⬇️ 匯出工單進度表（Excel）",
        data=buf,
        file_name=f"工單進度表_{date_start.strftime('%Y%m%d')}_{date_end.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
