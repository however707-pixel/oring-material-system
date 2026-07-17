# -*- coding: utf-8 -*-
"""工單紀錄.xlsx（NAS 改機排程）共用解析＋寫回邏輯

供 pages/22_pmc_order_tracking.py 與 scripts/send_no_wo_mail.py 共用，
規則改這裡即可，頁面與郵件報表同步生效。
"""
import io
import os
import shutil
from datetime import datetime

import pandas as pd

# 可用環境變數覆蓋（測試用）
WO_REC_PATH = os.environ.get(
    "PMC_WO_REC_PATH",
    r"\\192.168.2.34\MO_Storage\ORing MO\ORing-MO 工作\生管部\09. 廠內改機排程\工單紀錄.xlsx",
)
SHEET = "工作表1"
BACKUP_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                          "data", "backups")

KEEP_COLS = ["通知日期", "訂單日期", "訂單單號", "客戶簡稱", "品號", "品名", "訂單數量",
             "預交日", "成品工單號碼", "完工日", "開工日(組包)", "完工日(打件)", "開工日(打件)",
             "代工廠", "需求備註"]
KEEP_SET = set(KEEP_COLS)
CANON = {"目標入庫日": "完工日", "組包開工日": "開工日(組包)", "打件開工日": "開工日(打件)",
         "訂單回覆備註": "需求備註"}
NO_FIN = "未填完工日"


def clean_str(s: pd.Series) -> pd.Series:
    """空值安全的字串化（pandas 3 的 astype(str) 會保留 NaN，不能直接用）"""
    return (s.fillna("").astype(str).str.strip()
            .replace({"nan": "", "NaT": "", "None": ""}))


def fmt_cell(v):
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(v, pd.Timestamp) or hasattr(v, "strftime"):
        try:
            return pd.Timestamp(v).strftime("%Y/%m/%d")
        except (TypeError, ValueError):
            pass
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    return "" if s in ("nan", "NaT", "None") else s


def parse_workbook_full(file_bytes: bytes):
    """逐區塊解析多段表頭 log：以「通知日期」表頭列切塊，依各塊表頭對欄後合併

    回傳 (df, blocks_meta)：df 內含 _xlrow（工作表原始列，0-based）與 _xlblock（區塊序號）；
    blocks_meta[i] = {欄名: 工作表欄位索引(0-based)}，供寫回 Excel 定位。
    """
    try:
        raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=SHEET, header=None, engine="calamine")
    except Exception:
        raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=SHEET, header=None, engine="openpyxl")

    hdr_idx = raw.index[raw.iloc[:, 0].astype(str).str.strip() == "通知日期"].tolist()
    blocks, metas = [], []
    for i, h in enumerate(hdr_idx):
        end = hdr_idx[i + 1] if i + 1 < len(hdr_idx) else len(raw)
        body = raw.iloc[h + 1:end]
        if body.empty:
            continue
        header = raw.iloc[h]
        cand = {}
        for ci, v in header.items():
            if pd.isna(v):
                continue
            name = CANON.get(str(v).strip(), str(v).strip())
            if name in KEEP_SET:
                cand.setdefault(name, []).append(ci)
        # 同名欄位取資料較多的那一欄（處理重複表頭欄）
        colmap = {max(cis, key=lambda ci: int(body[ci].notna().sum())): name
                  for name, cis in cand.items()}
        sub = body[list(colmap)].rename(columns=colmap).reindex(columns=KEEP_COLS)
        sub["通知日期"] = sub["通知日期"].ffill()
        sub["_xlrow"] = sub.index
        sub["_xlblock"] = len(metas)
        metas.append({name: ci for ci, name in colmap.items()})
        blocks.append(sub)

    if not blocks:
        return pd.DataFrame(columns=KEEP_COLS), []
    df = pd.concat(blocks, ignore_index=True)

    # 去掉識別欄全空的雜訊列
    ids = df[["訂單單號", "品號", "品名", "成品工單號碼"]].apply(clean_str)
    df = df[ids.ne("").any(axis=1)].copy()

    df["_dt_通知"] = pd.to_datetime(df["通知日期"], errors="coerce")
    df["_dt_完工"] = pd.to_datetime(df["完工日"], errors="coerce")
    df["_qty"] = pd.to_numeric(df["訂單數量"], errors="coerce")
    return df, metas


def parse_workbook(file_bytes: bytes) -> pd.DataFrame:
    return parse_workbook_full(file_bytes)[0]


def enrich(df: pd.DataFrame, today) -> pd.DataFrame:
    """今年篩選＋分組（國智/其他）＋工單有無＋完工月份＋穩定列鍵，依完工日排序"""
    df = df[df["_dt_通知"] >= pd.Timestamp(today.year, 1, 1)].copy()
    if df.empty:
        return df

    df["_has_wo"] = clean_str(df["成品工單號碼"]).ne("")
    vend = clean_str(df["代工廠"])
    df["_grp"] = vend.str.contains("國智", na=False).map({True: "國智", False: "其他"})
    df["_month"] = df["_dt_完工"].dt.strftime("%Y/%m").fillna(NO_FIN)
    df = df.sort_values(["_dt_完工", "_dt_通知"], na_position="last").reset_index(drop=True)

    base = (df["_dt_通知"].dt.strftime("%Y%m%d").fillna("na") + "|"
            + clean_str(df["訂單單號"]) + "|" + clean_str(df["品號"]))
    df["_key"] = base + "#" + base.groupby(base).cumcount().astype(str)
    return df


# ─── 寫回 Excel（介面填寫 → NAS 原檔同步）────────────────────────────────────

DATE_WRITE_COLS = {"預交日", "完工日", "開工日(組包)", "完工日(打件)", "開工日(打件)"}


def _to_cell_value(col: str, text: str):
    """填寫文字 → Excel 儲存格值：日期欄能解析成日期就寫日期，其餘寫文字"""
    if col in DATE_WRITE_COLS:
        dt = pd.to_datetime(text, errors="coerce")
        if pd.notna(dt):
            return dt.to_pydatetime(), "yyyy/m/d"
    return text, None


def _backup_workbook():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    dst = os.path.join(BACKUP_DIR, f"工單紀錄_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(WO_REC_PATH, dst)
    # 只留最近 20 份備份
    backups = sorted(
        f for f in os.listdir(BACKUP_DIR)
        if f.startswith("工單紀錄_") and f.endswith(".xlsx")
    )
    for old in backups[:-20]:
        try:
            os.remove(os.path.join(BACKUP_DIR, old))
        except OSError:
            pass
    return dst


def sync_edits_to_excel(edits: dict, today) -> dict:
    """把介面填寫寫回 NAS 工單紀錄.xlsx。

    - 依當下檔案重新解析定位（檔案被別人改過也能對到正確列）
    - 逐列身分核對（訂單單號＋品號與列鍵一致才寫入）
    - 寫入前自動備份到 data/backups/
    - 檔案被鎖定（他人開啟中）時拋出例外，由呼叫端保留待同步

    回傳 {"synced": [key...], "skipped": {key: 原因}, "backup": 備份路徑或 None}
    """
    from openpyxl import load_workbook

    with open(WO_REC_PATH, "rb") as f:
        df, blocks = parse_workbook_full(f.read())
    df = enrich(df, today)
    loc = {k: (int(x), int(b)) for k, x, b in zip(df["_key"], df["_xlrow"], df["_xlblock"])}

    wb = load_workbook(WO_REC_PATH)
    ws = wb[SHEET]
    synced, skipped = [], {}

    for key, fields in edits.items():
        pos = loc.get(key)
        if pos is None:
            skipped[key] = "在最新檔案中找不到對應列"
            continue
        xlrow, bi = pos
        colmap = blocks[bi]

        # 身分核對：該列的 訂單單號、品號 必須與列鍵一致
        parts = key.split("|")
        oid = parts[1] if len(parts) > 1 else ""
        pno = parts[2].rsplit("#", 1)[0] if len(parts) > 2 else ""

        def cell_text(cname):
            ci = colmap.get(cname)
            if ci is None:
                return None
            v = ws.cell(row=xlrow + 1, column=ci + 1).value
            return "" if v is None else str(v).strip()

        if not oid and not pno:
            skipped[key] = "此列無單號與品號，無法安全定位"
            continue
        if pno and (cell_text("品號") or "") != pno:
            skipped[key] = "列位置核對不符（品號不一致），未寫入"
            continue
        if oid and (cell_text("訂單單號") or "") != oid:
            skipped[key] = "列位置核對不符（訂單單號不一致），未寫入"
            continue

        missing = [c for c in fields if colmap.get(c) is None]
        if missing:
            skipped[key] = f"該區塊缺少欄位 {missing}，未寫入"
            continue

        for col, text in fields.items():
            val, numfmt = _to_cell_value(col, str(text))
            c = ws.cell(row=xlrow + 1, column=colmap[col] + 1)
            c.value = val
            if numfmt:
                c.number_format = numfmt
        synced.append(key)

    backup = None
    if synced:
        backup = _backup_workbook()
        wb.save(WO_REC_PATH)  # 檔案被鎖定時在此拋 PermissionError
    return {"synced": synced, "skipped": skipped, "backup": backup}
